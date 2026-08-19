# -*- coding: utf-8 -*-
"""나라장터 채권평가회사 공고 수집기

narajangteo.py + 정밀수집.py + 목록만_추출.py 를 하나로 합친 것.
점수 판정은 점수.py / 점수표.yaml 이 맡는다.

동작 (요구사항_v1.md 의 1~6차)
  1차  면허제한 API 로 업종코드 3865 확인 → 있으면 무조건 수집        [A]
  2차  공고명 점수 (점수표.yaml)  13점↑ 수집후보 / 7~12점 검토 필요
  3차  2차 통과분만 첨부 PDF 를 열어 자격요건 확인
  4차  PDF 에 '채권평가회사' '3865' 등 발견 → 무조건 수집             [A]
  5차  PDF 키워드 없지만 13점 이상 → 수집                            [B]
  6차  PDF 키워드 없고 7~12점 → 같은 엑셀에 검토 필요로 표시          [C]

정확도 A/B/C 는 1·4차(키워드 일치) / 5차(일부 일치) / 6차(검토 필요) 다.

결과물은 엑셀 하나와 공고문 PDF 뿐이다 (요구사항_v2·v3).
애매한 건(C)을 따로 파일로 빼지 않고 공고목록.xlsx 에 함께 넣는다
(요구사항_v3 회신 3번). 엑셀에 오른 건의 공고문 PDF 만 `공고문/` 에 남기고,
그 밖의 첨부(HWP·ZIP 등)는 받지 않는다.

수집이 끝나면 그 회차 결과를 메일 본문에 표로 넣어 보낸다
(요구사항_v3 회신 1·2번). config.yaml 의 mail 섹션에서 켠다.

실행
  python 수집기.py              config.yaml 의 hours 만큼 (매일 자동실행용)
  python 수집기.py 30           최근 30일
  python 수집기.py 30 --목록만  PDF 저장 없이 엑셀만 (점수표 다듬을 때)
  python 수집기.py --pdf없이    PDF 검사 생략 (빠름, 1·2차만)
  python 수집기.py --메일없이   메일 발송만 건너뛴다
  python 수집기.py --메일만     수집 없이 엑셀의 마지막 회차를 메일로 (발송 시험용)

  지난 회차 재현 (그날 그 시각에 돌렸다면 어떤 메일이 나갔을지)
  python 수집기.py 0.75 --기준 "2026-08-06 10:00" --저장 ./재현_8월6일
"""

from __future__ import annotations

import csv
import io
import json
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import unquote

try:
    import requests
    import yaml
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("필요한 패키지를 설치하세요:\n\n    pip install requests pyyaml openpyxl\n")

try:
    from pypdf import PdfReader
    HAS_PDF = True
except ImportError:
    HAS_PDF = False


sys.path.insert(0, str(Path(__file__).resolve().parent))
from 앱경로 import FROZEN, app_dir  # noqa: E402

HERE = app_dir()
if not FROZEN:
    sys.path.insert(0, str(HERE))
from 점수 import Scorer  # noqa: E402

KST = timezone(timedelta(hours=9))
CONFIG_PATH = HERE / "config.yaml"

BASE = "https://apis.data.go.kr/1230000/ad/BidPublicInfoService"
ENDPOINTS = {
    "용역": f"{BASE}/getBidPblancListInfoServcPPSSrch",
    "물품": f"{BASE}/getBidPblancListInfoThngPPSSrch",
    "공사": f"{BASE}/getBidPblancListInfoCnstwkPPSSrch",
}
LICENSE_ENDPOINT = f"{BASE}/getBidPblancListInfoLicenseLimit"

CHUNK_DAYS = 14          # API 조회기간 제한
TIMEOUT = 90             # 면허제한은 응답이 크다. 넉넉히 잡는다
MAX_RETRY = 3
PAGE_SIZE = 500          # 100 이면 호출 수가 5배로 늘어 잘 끊긴다
PDF_MAX_MB = 20
PDF_MAX_PAGES = 12       # 자격요건은 대개 앞쪽에 있다

COLLECT_XLSX = "공고목록.xlsx"
OLD_REVIEW_XLSX = "검토후보.xlsx"   # 한 파일로 합치기 전에 쓰던 파일 (안내용)
SEEN_NAME = "_수집이력.json"
PDF_DIR = "공고문"

HEADERS = ["정확도", "공고종류", "변경내역", "등록일", "수요기관", "공고기관",
           "공고명", "마감일시", "배정예산", "추정가격", "업무", "점수", "키워드",
           "근거", "첨부", "출처", "공고번호", "링크", "수집일시"]
OLD_GRADE = {"S": "A"}
WIDTHS = [6, 8, 34, 12, 24, 24, 58, 17, 14, 14, 6, 6, 16, 40, 6, 10, 18, 12, 17]
COL_GRADE = 0
COL_KIND = 1             # 공고종류 (신규/변경/재공고/취소)
COL_DIFF = 2             # 변경내역 (우리 이력과 견줘 달라진 것. 신규는 빈칸)
COL_REG = 3              # 등록일
COL_DM = 4               # 수요기관
COL_NT = 5               # 공고기관
COL_TITLE = 6
COL_DUE = 7              # 마감일시
COL_BUDGET = 8           # 배정예산 (부가세 포함)
COL_PRICE = 9            # 추정가격 (부가세 제외)
COL_SCORE = 11
COL_KW = 12              # 키워드
COL_WHY = 13             # 근거
COL_SRC = 15             # 출처 (지금은 모두 '나라장터')
COL_NO = 16              # 공고번호
COL_LINK = 17            # 링크 (하이퍼링크로 바꾸는 칸)
COL_STAMP = 18           # 수집일시 (한 회차는 값이 모두 같다)

# 지금은 나라장터 하나뿐이다. 기관 자체 홈페이지까지 보게 되면 여기가 늘고,
# 수집이력도 출처마다 따로 세어야 한다(이력키 참고). 그때 컬럼을 새로 넣으면
# 이미 쌓인 엑셀을 다 옮겨야 해서 지금 한 칸 잡아둔다.
SRC_G2B = "나라장터"

# 컬럼 구성은 몇 번 바뀌었고 앞으로도 바뀐다. --메일만 은 그 전에 쌓아둔
# 엑셀도 읽을 수 있어야 해서, 자리(몇 번째 칸)가 아니라 첫 줄에 적힌 이름을
# 보고 지금 구성의 어느 칸인지 찾는다 (read_saved_rows 참고).
# 이름이 바뀐 칸만 여기 적어둔다.  옛이름 → 지금이름
RENAMED = {"등급": "정확도"}

# 정확도 3단계. A 는 근거가 두 가지(면허제한 / 공고문)지만 결론이 같아 합쳤다.
GRADE_DESC = {
    "A": "시가/대체 평가 및 채권평가회사(업종코드) 키워드 일치",
    "B": "키워드 일부 일치",
    "C": "기금/공제회 관련, 검토 필요 공고",
}
# 메일 본문에서는 알파벳 한 글자만으로는 뜻이 안 통한다.
GRADE_LABEL = {"A": "A 일치", "B": "B 일부", "C": "C 검토"}

# 메일 문구의 기본값. config.yaml 의 mail.문구 로 덮어쓸 수 있다.
# 담당자가 바뀌어도 파이썬을 안 열고 고칠 수 있어야 해서 밖으로 뺐다.
# {중괄호} 는 실행할 때 값으로 바뀐다. 없는 이름을 쓰면 그 줄만 기본값으로
# 되돌리고 화면에 알린다 (fill 참고). 메일이 안 나가는 일은 없다.
MAIL_TEXT = {
    "제목": "[나라장터] {날짜} 신규 {건수}건 ({내역})",
    "제목_없음": "[나라장터] {날짜} 입찰 공고 특이사항 없습니다",
    "제목_변경": "[나라장터] {날짜} 공고 {건수}건 ({내역})",
    "제목_일부": "[나라장터] {날짜} 공고 {건수}건 ({내역}) — 일부 조회 실패",
    "제목_실패": "[나라장터] {날짜} 자동수집 실패 — 확인 필요",
    "실패알림": ("이번 회차는 조회가 완전하지 않았습니다. 아래 목록에 빠진 "
                 "공고가 있을 수 있습니다.\n"
                 "못 받은 구간: {빠짐}"),
    "첫줄": "{시각} 기준 나라장터 신규 공고입니다. ({기간})",
    "요약": "총 {건수}건 — {내역}",
    "변경머리": "■ 이미 알려드린 공고 중 바뀐 것 {건수}건",
    "신규머리": "■ 새로 올라온 공고 {건수}건",
    "없음": "{날짜} 나라장터 입찰 공고 특이사항 없습니다.",
    "안내": ("공고명을 누르면 나라장터 공고 화면이 열립니다.\n"
             "A = 시가/대체 평가 및 채권평가회사(업종코드) 키워드 일치 · "
             "B = 키워드 일부 일치 · "
             "C = 기금/공제회 관련, 검토 필요 공고 (회색 줄)"),
    "검토안내": ("C는 자격요건에 채권평가회사 표기가 없어 판정이 애매한 "
                 "건입니다. 공고문을 열어 확인이 필요합니다."),
    "꼬리말": "자동 발송 메일입니다. 원본 파일 위치: {폴더}",
}
GRADE_COLOR = {"A": "C00000", "B": "1F3864", "C": "808080"}
GRADE_ORDER = {"A": 0, "B": 1, "C": 2}
# C(검토 필요)는 같은 파일에 들어가므로 행 전체를 연회색으로 깔아 구분한다.
GRADE_FILL = {"C": "F2F2F2"}

# 공고종류. 왼쪽이 API 의 ntceKindNm 원문, 오른쪽이 엑셀·메일에 적는 말이다.
# 취소공고는 지금 조회 단계에서 버리므로(pick_live) 엑셀까지 오지 않지만,
# 나중에 취소도 알리기로 하면 쓸 자리라 미리 적어둔다.
NOTICE_KIND = {"등록공고": "신규", "변경공고": "변경",
               "재공고": "재공고", "취소공고": "취소"}
# 신규가 대부분이라 그대로 두면 눈에 안 띈다. 변경·재공고만 색을 준다.
KIND_COLOR = {"변경": "#C00000", "재공고": "#BF8F00"}

BAD_CHARS = re.compile(r'[\\/:*?"<>|\r\n\t]')
MAX_PATH = 250           # 윈도우 260자 제한에서 여유분을 뺀 값

# 한 번 실행하면 그 회차 전 행에 같은 값이 들어간다.
# 행마다 초 단위로 다르면 '언제 돌린 배치인지' 로 묶어보기가 어렵다.
# --기준 으로 지난 회차를 재현할 때는 그 시각으로 바뀐다.
RUN_STAMP = f"{datetime.now(KST):%Y-%m-%d %H:%M}"


LOG_DIR = HERE / "로그"
LOG_MAX_BYTES = 2_000_000   # 넘으면 실행.log.1 로 밀어내고 새로 쓴다.

# 수집기를 직접 실행할 때만 기록한다.
# 진단·실측 스크립트가 import 해서 쓰는 경우까지 남기면 로그가 지저분해진다.
# exe 로 묶으면 진입점이 실행진입.py 라 __name__ 이 "__main__" 이 아니다.
# 그쪽에서 enable_log() 를 불러 준다.
_LOG_ENABLED = __name__ == "__main__"
_LOG_FILE = None


def enable_log() -> None:
    """실행 기록을 남기게 한다. 수집을 진짜로 돌리는 진입점만 부른다."""
    global _LOG_ENABLED
    _LOG_ENABLED = True


def _log_file():
    """실행 기록을 남길 파일. 처음 부를 때 열고 헤더를 적는다.

    화면만 보고 있으면 작업 스케줄러가 새벽에 왜 실패했는지 알 수 없다.
    기록에 실패해도 수집 자체는 계속되어야 하므로 오류는 삼킨다.
    """
    global _LOG_FILE, _LOG_ENABLED
    if _LOG_FILE is not None or not _LOG_ENABLED:
        return _LOG_FILE
    try:
        LOG_DIR.mkdir(exist_ok=True)
        path = LOG_DIR / "실행.log"
        if path.exists() and path.stat().st_size > LOG_MAX_BYTES:
            path.replace(LOG_DIR / "실행.log.1")
        _LOG_FILE = open(path, "a", encoding="utf-8")
        옵션 = " ".join(sys.argv[1:]) or "(옵션 없음)"
        _LOG_FILE.write(f"\n==== 실행 {datetime.now(KST):%Y-%m-%d %H:%M:%S}"
                        f"  {옵션} ====\n")
        _LOG_FILE.flush()
    except OSError:
        _LOG_ENABLED = False     # 다시 시도하지 않는다.
    return _LOG_FILE


def log(msg: str = "") -> None:
    print(msg, flush=True)
    f = _log_file()
    if f is not None:
        try:
            f.write(f"{msg}\n")
            f.flush()
        except OSError:
            pass


# ---------------------------------------------------------------------------
# 설정
# ---------------------------------------------------------------------------


def normalize_key(key: str) -> str:
    """Encoding 키를 넣어도 동작하게 한다.

    공공데이터포털은 같은 키를 두 형태로 준다.
      Decoding : abc+de/fg==      ← 이게 정답
      Encoding : abc%2Bde%2Ffg%3D%3D
    requests 가 파라미터를 다시 인코딩하므로 Encoding 키를 그대로 넣으면
    %2F 가 %252F 로 이중 인코딩돼 403이 난다.
    """
    if not key:
        return key
    if re.search(r"%2[BbFf]|%3[Dd]", key):
        log("[안내] Encoding 인증키가 감지되어 자동으로 변환했습니다.")
        log("       원래는 마이페이지의 'Decoding' 키를 넣어야 합니다.")
        return unquote(key)
    return key


def resolve_output_dir(raw: str) -> Path:
    """저장 폴더를 확정한다. 네트워크 경로(UNC)도 그대로 지원한다."""
    raw = raw.strip().strip('"')
    out = Path(raw)
    if not out.is_absolute():
        out = HERE / out

    text = str(out)
    is_unc = text.startswith("\\\\") or text.startswith("//")

    # 매핑 드라이브(Z: 등)는 로그인 세션에만 존재한다.
    # 작업 스케줄러가 '로그온 여부 관계없이 실행'으로 돌면 드라이브를 못 찾는다.
    if re.match(r"^[A-Za-z]:", text) and text[0].upper() not in "CD":
        log(f"[안내] 저장 위치가 드라이브 문자({text[:2]})입니다.")
        log("       작업 스케줄러로 자동 실행할 계획이면 UNC 경로를 쓰세요.")
        log(r"       예:  output_dir: '\\서버이름\공유폴더\조달'")

    try:
        out.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        hint = ""
        if is_unc:
            hint = ("\n  네트워크 폴더에 접근할 수 없습니다. 탐색기에서 해당 경로가"
                    "\n  열리는지, 쓰기 권한이 있는지 먼저 확인하세요.")
        sys.exit(f"저장 폴더를 만들 수 없습니다:\n  {out}\n  → {exc}{hint}")

    probe = out / ".쓰기테스트"
    try:
        probe.write_text("ok", encoding="utf-8")
        probe.unlink()
    except OSError:
        sys.exit(f"저장 폴더에 쓸 수 없습니다 (권한 확인 필요):\n  {out}")

    if is_unc:
        log(f"[안내] 네트워크 폴더에 저장합니다: {out}")
    return out


def load_config() -> dict:
    if not CONFIG_PATH.exists():
        sys.exit(f"설정 파일이 없습니다: {CONFIG_PATH}")

    with CONFIG_PATH.open(encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}

    key = normalize_key(str(cfg.get("service_key", "")).strip().strip("'\""))
    if not key or key.startswith("여기에"):
        sys.exit(
            "config.yaml 의 service_key 가 비어 있습니다.\n"
            "공공데이터포털 마이페이지에서 'Decoding' 인증키를 복사해 넣으세요."
        )
    cfg["service_key"] = key

    cfg["hours"] = int(cfg.get("hours") or 24)
    cfg["targets"] = [t for t in (cfg.get("targets") or ["용역"]) if t in ENDPOINTS]
    if not cfg["targets"]:
        sys.exit(f"targets 가 잘못됐습니다. 사용 가능: {', '.join(ENDPOINTS)}")
    cfg["industry_codes"] = [str(c).strip()
                             for c in (cfg.get("industry_codes") or ["3865"])]
    cfg["output_dir"] = resolve_output_dir(str(cfg.get("output_dir") or "./수집결과"))
    cfg["mail"] = load_mail_config(cfg.get("mail"))
    return cfg


def as_list(value) -> list[str]:
    """`to: a@b.com` 처럼 한 줄로 써도, 목록으로 써도 받는다."""
    if not value:
        return []
    if isinstance(value, str):
        return [x.strip() for x in re.split(r"[,;]", value) if x.strip()]
    return [str(x).strip() for x in value if str(x).strip()]


def load_mail_config(raw) -> dict:
    """메일 설정을 확정한다. 주소가 비어 있으면 조용히 끈다.

    설정이 덜 채워진 상태로 자동 실행되는 일이 흔하다. 그때 예외로 죽으면
    수집 결과까지 날아가므로, 못 보내는 상황이면 수집만 하고 넘어간다.
    """
    m = dict(raw or {})
    mail = {
        "enabled": bool(m.get("enabled", False)),
        "mode": str(m.get("mode") or "outlook").strip().lower(),
        "to": as_list(m.get("to")),
        "cc": as_list(m.get("cc")),
        # 오류 메일 받는 사람. 업무 담당자에게 API 오류 메일이 가면 안 읽는다.
        # 비워두면 to 로 간다 (아예 안 가는 것보다 낫다).
        "error_to": as_list(m.get("error_to")),
        "문구": {k: str(v) for k, v in (m.get("문구") or {}).items()
                 if v is not None and str(v).strip()},
        "draft_only": bool(m.get("draft_only", False)),
        "attach_excel": bool(m.get("attach_excel", True)),
        "attach_pdf": bool(m.get("attach_pdf", False)),
        "attach_max_mb": float(m.get("attach_max_mb") or 15),
        "send_when_empty": bool(m.get("send_when_empty", False)),
        "smtp": dict(m.get("smtp") or {}),
    }
    unknown = [k for k in mail["문구"] if k not in MAIL_TEXT]
    if unknown:
        log(f"[주의] mail.문구 에 모르는 항목이 있어 무시합니다: {', '.join(unknown)}")
        log(f"       쓸 수 있는 항목: {', '.join(MAIL_TEXT)}")
    if mail["mode"] not in ("outlook", "smtp"):
        log(f"[주의] mail.mode 는 outlook 또는 smtp 여야 합니다 (지금: {mail['mode']}).")
        mail["enabled"] = False
    if mail["enabled"] and not mail["to"]:
        log("[주의] mail.to 가 비어 있어 메일을 보내지 않습니다.")
        mail["enabled"] = False
    return mail


# ---------------------------------------------------------------------------
# API 조회
# ---------------------------------------------------------------------------

_session: requests.Session | None = None


def session() -> requests.Session:
    """연결을 재사용한다. 매 요청마다 TLS 핸드셰이크를 새로 하면 느리고 잘 끊긴다."""
    global _session
    if _session is None:
        _session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(pool_connections=4, pool_maxsize=4)
        _session.mount("https://", adapter)
        _session.mount("http://", adapter)
    return _session


def get_page(endpoint: str, params: dict, label: str) -> dict | None:
    """한 페이지를 가져온다. 일시적 실패는 재시도한다."""
    for attempt in range(1, MAX_RETRY + 1):
        try:
            res = session().get(endpoint, params=params, timeout=TIMEOUT)
            res.raise_for_status()
            return res.json()
        except (requests.Timeout, requests.ConnectionError) as exc:
            if attempt == MAX_RETRY:
                log(f"  [실패] {label} — {type(exc).__name__} ({MAX_RETRY}회 재시도 후 포기)")
                return None
            wait = 2 ** attempt
            log(f"  [재시도 {attempt}/{MAX_RETRY}] {label} — {wait}초 후 다시 시도")
            time.sleep(wait)
        except requests.RequestException as exc:
            log(f"  [실패] {label} — {exc}")
            if "403" in str(exc):
                log("         → 인증키(Decoding) 또는 활용신청 승인 상태를 확인하세요.")
            elif "429" in str(exc):
                log("         → 일일 호출 한도 초과입니다. 내일 다시 실행하세요.")
            return None
        except json.JSONDecodeError:
            log(f"  [실패] {label} — 응답이 JSON이 아닙니다. 인증키를 확인하세요.")
            return None
    return None


def call_api(endpoint: str, key: str, begin: datetime, end: datetime,
             label: str = "") -> tuple[list[dict], list[str]]:
    """(수집된 행, 못 받은 구간 목록) 반환.

    '완전한가' 만 알면 메일에 "일부 실패" 라고만 쓸 수 있다. 어느 날짜가
    빠졌는지 적어줘야 그 구간만 --기준 으로 다시 돌릴 수 있으므로,
    끊긴 구간을 '08/12 09:00~08/13 09:00' 형태로 모아 돌려준다.
    """
    rows: list[dict] = []
    missing: list[str] = []
    cur = begin

    while cur < end:
        chunk_end = min(cur + timedelta(days=CHUNK_DAYS), end)
        page = 1
        while True:
            params = {
                "serviceKey": key, "pageNo": page, "numOfRows": PAGE_SIZE,
                "inqryDiv": 1,
                "inqryBgnDt": cur.strftime("%Y%m%d%H%M"),
                "inqryEndDt": chunk_end.strftime("%Y%m%d%H%M"),
                "type": "json",
            }
            payload = get_page(endpoint, params, f"{label} {cur:%m/%d} p{page}")
            if payload is None:
                missing.append(f"{cur:%m/%d %H:%M}~{chunk_end:%m/%d %H:%M}")
                break

            # 한도초과·인증오류는 HTTP 200 에 'response' 가 아예 없는 형태로 온다.
            #   {"OpenAPI_ServiceResponse": {"cmmMsgHeader": {"errMsg": ...}}}
            # 이걸 걸러내지 않으면 resultCode 가 None 이 되어 아래 검사를 통과하고,
            # items 가 비어 그냥 break 하면서 그 구간이 '정상 0건' 으로 남는다.
            # 일일한도를 넘긴 날 "특이사항 없습니다" 메일이 나가는 경로가 이것이다.
            if "OpenAPI_ServiceResponse" in payload:
                msg = payload["OpenAPI_ServiceResponse"].get("cmmMsgHeader", {})
                reason = (msg.get("returnAuthMsg") or msg.get("errMsg")
                          or "알 수 없는 오류")
                log(f"  [실패] {label} — {reason}")
                missing.append(f"{cur:%m/%d %H:%M}~{chunk_end:%m/%d %H:%M}"
                               f" ({reason})")
                break

            header = payload.get("response", {}).get("header", {})
            code = header.get("resultCode")
            # None 을 허용하면 안 된다. 위에서 걸러지지 않은 낯선 응답 형태도
            # 여기서 실패로 잡아야 '0건' 으로 둔갑하지 않는다.
            if code not in ("00", "0"):
                log(f"  [오류] {label} API 응답 {code}: {header.get('resultMsg')}")
                missing.append(f"{cur:%m/%d %H:%M}~{chunk_end:%m/%d %H:%M}"
                               f" (응답 {code})")
                break

            body = payload.get("response", {}).get("body", {})
            items = body.get("items") or []
            if isinstance(items, dict):
                items = items.get("item", [])
            if not items:
                break

            rows.extend(items)
            total = int(body.get("totalCount", 0) or 0)
            if page * PAGE_SIZE >= total:
                break
            page += 1
        cur = chunk_end

    return rows, missing


# ---------------------------------------------------------------------------
# 1차 — 면허제한 (업종코드)
# ---------------------------------------------------------------------------

CODE_RE = re.compile(r"/\s*(\d{3,5})")


def parse_codes(text: str) -> set[str]:
    """'액화석유가스판매사업/4617' 이나 '[명칭1/4615],[명칭2/4616]' 에서 코드만 뽑는다."""
    return set(CODE_RE.findall(text or ""))


def merge_license_rows(index: dict, rows: list[dict]) -> None:
    """조회 결과를 {공고번호: (업종코드집합, 업종명목록)} 으로 누적한다.

    면허제한 API 는 허용 업종마다 한 줄씩 준다. 예를 들어
    '회계법인/1200', '종합금융투자사업자/7096', '채권평가회사/3865' 가
    세 줄로 온다. 업종명을 첫 줄만 남기면 정작 우리가 걸린 업종이
    안 보이므로 전부 모은다. 실제로 공고의 37% 가 두 줄 이상이다.

    차수는 키에 넣지 않는다. 면허제한이 차수 000 에만 실려 있는데
    공고 쪽은 변경공고라 차수가 001 인 경우가 있어, 차수까지 맞추면 못 찾는다.
    """
    for r in rows:
        uid = (r.get("bidNtceNo") or "").strip()
        limit_nm = (r.get("lcnsLmtNm") or "").strip()
        permit = (r.get("permsnIndstrytyList") or "").strip()
        codes = parse_codes(limit_nm) | parse_codes(permit)

        old_codes, labels = index.get(uid, (set(), []))
        for name in (limit_nm, permit):
            if name and name not in labels:
                labels.append(name)
        index[uid] = (old_codes | codes, labels)


def license_chunks(begin: datetime, end: datetime) -> list[tuple[datetime, datetime]]:
    """면허제한 조회를 14일 구간으로 나눈다.

    경계를 자정에 맞춘다. 실행 시각이 몇 분만 달라도 구간이 어긋나면
    캐시가 하나도 안 맞아 처음부터 다시 받게 되기 때문이다.
    맨 앞 구간이 조금 넓어지지만 면허제한은 공고번호로 찾아 쓰는 색인이라
    범위가 남는 것은 문제가 되지 않는다.
    """
    chunks = []
    cur = begin.replace(hour=0, minute=0, second=0, microsecond=0)
    while cur < end:
        nxt = cur + timedelta(days=CHUNK_DAYS)
        chunks.append((cur, min(nxt, end)))
        cur = nxt
    return chunks


def build_license_index(key: str, begin: datetime, end: datetime,
                        cache_dir: Path) -> tuple[dict, list[str]]:
    """({공고번호-차수: (업종코드집합, 표시용문자열)}, 못 받은 구간 목록)

    조회량이 많아 (4개월이면 300회가 넘는다) 중간에 끊기기 쉽다.
    그래서 14일 구간마다 따로 캐시에 남긴다. 끊겨도 받아둔 구간은 남으므로
    다시 실행하면 못 받은 구간부터 이어간다.

    아직 안 끝난 마지막 구간은 캐시하지 않는다. 그 구간은 시간이 지나면
    공고가 더 쌓이므로, 캐시해두면 오늘 새로 올라온 공고를 못 본다.
    """
    index: dict[str, tuple[set, str]] = {}
    missing: list[str] = []
    chunks = license_chunks(begin, end)

    for i, (c_begin, c_end) in enumerate(chunks, start=1):
        tag = f"{i}/{len(chunks)} {c_begin:%m/%d}~{c_end:%m/%d}"
        closed = c_end - c_begin >= timedelta(days=CHUNK_DAYS)
        cache = cache_dir / f"_면허제한_{c_begin:%Y%m%d}_{c_end:%Y%m%d}.json"

        if closed and cache.exists():
            try:
                saved = json.loads(cache.read_text(encoding="utf-8"))
                for k, v in saved["index"].items():
                    # 예전 캐시는 '공고번호-차수' 키에 업종명이 문자열 하나였다.
                    k = k.rpartition("-")[0] or k
                    old = v[1] if isinstance(v[1], list) else ([v[1]] if v[1] else [])
                    codes, labels = index.get(k, (set(), []))
                    index[k] = (codes | set(v[0]),
                                labels + [x for x in old if x not in labels])
                log(f"  [{tag}] 캐시 사용 {len(saved['index']):,}건")
                continue
            except (json.JSONDecodeError, KeyError, TypeError, OSError):
                cache.unlink(missing_ok=True)

        rows, 빠짐 = call_api(LICENSE_ENDPOINT, key, c_begin, c_end,
                             label=f"면허제한 {tag}")
        merge_license_rows(index, rows)
        log(f"  [{tag}] {len(rows):,}건{'' if not 빠짐 else '  ← 끊김'}")

        if 빠짐:
            missing += 빠짐
            continue
        if closed:
            chunk_index: dict[str, tuple[set, list]] = {}
            merge_license_rows(chunk_index, rows)
            try:
                cache.write_text(json.dumps(
                    {"complete": True,
                     "index": {k: [sorted(v[0]), v[1]]
                               for k, v in chunk_index.items()}},
                    ensure_ascii=False), encoding="utf-8")
            except OSError:
                pass

    return index, missing


# ---------------------------------------------------------------------------
# 조회 기록 — '신규 0건' 과 '자동수집 실패' 를 갈라 보기 위한 것
#
# 0건 메일과 실패 메일이 똑같이 생기면, API 가 죽은 날에도 받는 사람은
# '오늘은 공고가 없었구나' 로 읽는다. 그게 곧 입찰 누락이다.
#
# 특히 위험한 것은 완전 실패가 아니라 부분 실패다. 한 구간만 못 받아도
# 메일은 '오늘 3건' 이라고 멀쩡히 나가지만 사실은 '10건 중 3건' 일 수 있다.
# 그래서 조회마다 성공·실패와 못 받은 구간을 남기고, 그 결과로 회차 상태를
# 정상 / 일부실패 / 완전실패 셋으로 가른다.
# ---------------------------------------------------------------------------

CHECK_LOG = "_조회기록.csv"     # 회차마다 한 줄씩 쌓이는 성공·실패 기록

정상, 일부실패, 완전실패 = "정상", "일부실패", "완전실패"


def check(이름: str, 건수: int, 빠짐: list[str], 필수: bool = True) -> dict:
    """조회 한 건의 결과. 필수=False 는 실패해도 회차를 실패로 보지 않는다."""
    return {"이름": 이름, "건수": 건수, "빠짐": list(빠짐), "필수": 필수}


def run_status(checks: list[dict]) -> str:
    """회차 상태. 목록 조회가 전부 실패했으면 결과 자체를 믿을 수 없다."""
    목록 = [c for c in checks if c["필수"]]
    if 목록 and all(c["빠짐"] for c in 목록):
        return 완전실패
    return 일부실패 if any(c["빠짐"] for c in checks) else 정상


def missing_phrase(checks: list[dict]) -> str:
    """'용역 08/12 09:00~08/13 09:00 · 면허제한 08/11~08/12' 처럼 만든다."""
    조각 = []
    for c in checks:
        for 구간 in c["빠짐"]:
            조각.append(f"{c['이름']} {구간}")
    return " · ".join(조각)


def save_check_log(root: Path, checks: list[dict], status: str,
                   수집건수: int, begin: datetime, end: datetime) -> None:
    """회차별 조회 성공·실패를 한 파일에 쌓는다.

    실행.log 는 회차마다 수십 줄이라 '지난 두 주 동안 몇 번 실패했나' 를
    볼 수 없다. 한 줄짜리 기록이 따로 있어야 한다.
    엑셀에서 바로 열리도록 BOM 을 붙인다.
    """
    path = root / CHECK_LOG
    첫줄 = not path.exists()
    try:
        with open(path, "a", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            if 첫줄:
                w.writerow(["실행일시", "조회기간", "상태", "조회",
                            "건수", "못받은구간", "수집건수"])
            for c in checks:
                w.writerow([RUN_STAMP,
                            f"{begin:%Y-%m-%d %H:%M}~{end:%Y-%m-%d %H:%M}",
                            status, c["이름"], c["건수"],
                            " · ".join(c["빠짐"]), 수집건수])
    except OSError as exc:
        log(f"[주의] 조회기록을 남기지 못했습니다: {exc}")


def rerun_command(begin: datetime, end: datetime) -> str:
    """못 받은 구간을 다시 훑는 명령. 오류 메일에 그대로 넣는다.

    조회 기간은 실행할 때마다 앞으로 밀린다. 오늘 놓친 공고는 내일 그냥
    다시 돌려도 안 잡히므로, 그 회차를 재현하는 명령을 알려줘야 한다.
    """
    일수 = max(1, round((end - begin).total_seconds() / 86400))
    앞 = "나라장터수집기.exe 수집" if FROZEN else "python 수집기.py"
    return f'{앞} {일수} --기준 "{end:%Y-%m-%d %H:%M}"'


# ---------------------------------------------------------------------------
# 살아있는 공고 고르기
# ---------------------------------------------------------------------------


def deadline(raw: dict) -> str:
    """마감일시. 없으면 개찰일시로 대신하고 그렇다고 표시한다.

    직찰·우편 방식 공고는 전자입찰 마감시각(bidClseDt)이 아예 없다.
    나라장터 화면에서도 입찰마감일시가 공란으로 나온다.
    08/04 하루치 용역 624건 중 90건(14%)이 그랬고, 그 90건은 모두
    개찰일시(opengDt)가 채워져 있었다. 빈칸으로 두는 것보다 낫다.
    """
    clse = (raw.get("bidClseDt") or "").strip()
    if clse:
        return clse[:16]
    opng = (raw.get("opengDt") or "").strip()
    return f"{opng[:16]} (개찰)" if opng else ""


def notice_kind(raw: dict) -> str:
    """공고종류. API 의 ntceKindNm 을 화면에서 쓰는 말로 바꾼다.

    8/12~8/18 용역 1,500건을 세어보니 등록공고 1,054 · 재공고 285 ·
    변경공고 87 · 취소공고 74 였다. 즉 신규인지 변경인지 재공고인지는
    지난 수집분과 대조할 것 없이 API 가 그냥 알려준다.

    다만 지금은 수집이력에 있는 공고번호를 건너뛰므로(run 참고), 여기 '변경'
    으로 찍히는 건은 '우리가 처음 보는데 이미 한 번 바뀐 공고' 다.
    이미 보낸 공고가 나중에 바뀐 것을 잡으려면 이력에 값을 남겨야 한다.
    """
    return NOTICE_KIND.get((raw.get("ntceKindNm") or "").strip(), "-")


def money(raw) -> int | str:
    """예산 칸. 숫자로 넣어야 엑셀에서 정렬·합계가 된다.

    1,500건 중 6건은 금액이 비어 있거나 0 이었다. 그때는 0 을 넣지 않고
    빈칸으로 둔다. 0 을 넣으면 '예산 0원짜리 공고' 로 읽힌다.
    """
    text = str(raw or "").strip().replace(",", "")
    if not text or text == "0":
        return ""
    try:
        return int(float(text))
    except ValueError:
        return ""


def no_of(raw: dict) -> str:
    """공고번호. 여러 곳에서 같은 방식으로 꺼내야 해서 하나로 묶는다."""
    return (raw.get("bidNtceNo") or "").strip()


def notice_ord(raw: dict) -> int:
    try:
        return int(str(raw.get("bidNtceOrd") or "0").strip() or 0)
    except ValueError:
        return 0


def pick_live(raws: list[dict]) -> tuple[list[dict], dict, int]:
    """(살아있는 공고, 취소된 공고 {번호: 공고}, 옛 차수로 버린 수)

    목록 API 는 같은 공고를 차수마다 한 줄씩 준다. 차수는 재공고·변경·취소가
    생길 때마다 올라가므로(API 문서: '증가되는 수') 가장 큰 차수가 최신이다.
    나라장터 화면은 최신 차수만 보여주는데, 그대로 두면 엑셀에는 한 공고가
    여러 줄로 쌓인다.

    취소공고는 살아있는 목록에서 뺀다. 취소 줄만 빼면 이미 없어진 원공고가
    마감일시까지 달고 남아 살아있는 공고처럼 보이기 때문이다. 실제로 공고를
    취소하고 몇 분 뒤 새 번호로 다시 올리는 기관이 있다.

    다만 버리지는 않고 따로 돌려준다. 우리가 이미 메일로 보낸 공고가
    취소됐다면 그것도 알려야 한다. 조용히 사라지면 담당자는 그 공고가
    아직 살아있는 줄 안다.
    """
    최신: dict[str, dict] = {}
    취소: dict[str, dict] = {}

    for r in raws:
        no = (r.get("bidNtceNo") or "").strip()
        if not no:
            continue
        if (r.get("ntceKindNm") or "").strip() == "취소공고":
            취소[no] = r
        cur = 최신.get(no)
        if cur is None or notice_ord(r) > notice_ord(cur):
            최신[no] = r

    live = [r for no, r in 최신.items() if no not in 취소]
    return live, 취소, len(raws) - len(최신)


# ---------------------------------------------------------------------------
# 이미 보낸 공고가 달라졌는가
#
# API 는 '이 공고가 변경공고다' 까지만 알려주고 무엇이 바뀌었는지는 말해주지
# 않는다. 변경일시(chgDt)는 1,500건 중 1건만 채워져 있어 쓸 수 없다.
# 그래서 우리가 지난 회차에 적어둔 값과 직접 견준다.
# ---------------------------------------------------------------------------


def diff_notice(이전: dict, raw: dict, 차수: int) -> str:
    """지난번과 달라진 것을 사람이 읽는 한 줄로. 같으면 빈 문자열.

    이전 이 비어 있으면(옛 이력에서 넘어온 항목) 비교할 수 없으므로 빈
    문자열을 준다. 모르는 것을 '변경' 이라고 말하면 안 된다.
    """
    if not 이전:
        return ""

    조각 = []
    옛마감, 새마감 = str(이전.get("마감") or ""), deadline(raw)
    if 옛마감 and 새마감 and 옛마감 != 새마감:
        조각.append(f"마감 {옛마감} → {새마감}")

    옛예산, 새예산 = 이전.get("예산"), money(raw.get("asignBdgtAmt"))
    if isinstance(옛예산, int) and isinstance(새예산, int) and 옛예산 != 새예산:
        조각.append(f"예산 {옛예산:,} → {새예산:,}")

    옛이름 = str(이전.get("공고명") or "")
    새이름 = (raw.get("bidNtceNm") or "").strip()
    if 옛이름 and 새이름 and 옛이름 != 새이름:
        조각.append("공고명 바뀜")

    if 조각:
        return " · ".join(조각)

    # 눈에 보이는 값은 그대로인데 차수만 올라간 경우. 규격서나 첨부만 고친
    # 것이라 목록 API 로는 무엇이 바뀌었는지 알 수 없다. 그래도 알려준다.
    # 공고문을 다시 열어봐야 하는 건이기 때문이다.
    옛차수 = 이전.get("차수")
    if isinstance(옛차수, int) and 차수 > 옛차수:
        return f"재게시 (차수 {옛차수:03d} → {차수:03d}, 목록에 보이는 값은 그대로)"
    return ""


def cancel_row(이전: dict, raw: dict, 출처: str = SRC_G2B) -> list:
    """취소된 공고 한 줄. 판정을 새로 하지 않고 지난번 값을 그대로 쓴다.

    이미 한 번 메일에 실어 보낸 공고다. 지금 다시 채점해서 등급이 달라지면
    받는 사람만 헷갈린다. 알려야 할 것은 '그때 그 공고가 취소됐다' 뿐이다.
    """
    return [이전.get("정확도") or "C", "취소", "공고 취소됨",
            (raw.get("bidNtceDt") or "")[:10],
            (raw.get("dminsttNm") or "").strip(),
            (raw.get("ntceInsttNm") or "").strip(),
            (raw.get("bidNtceNm") or "").strip() or 이전.get("공고명") or "",
            deadline(raw), money(raw.get("asignBdgtAmt")),
            money(raw.get("presmptPrce")), "", 이전.get("점수") or 0,
            "-", "이미 보낸 공고가 취소됐습니다", 0, 출처,
            (raw.get("bidNtceNo") or "").strip(),
            (raw.get("bidNtceDtlUrl") or "").strip(), RUN_STAMP]


# ---------------------------------------------------------------------------
# 3·4차 — 첨부 PDF 에서 입찰참가자격 확인
# ---------------------------------------------------------------------------

QUAL_HEAD = re.compile(r"입\s*찰\s*참\s*가\s*자?\s*격|참가자격|자격요건|참가자격요건")
QUAL_TAIL = re.compile(r"입찰참가서류|제출서류|입찰보증금|낙찰자\s*결정|입찰방법|제출방법")


def fetch_pdf(url: str) -> bytes | None:
    """PDF 첨부를 메모리로 내려받는다. PDF 가 아니거나 실패하면 None."""
    try:
        res = session().get(url, timeout=TIMEOUT)
        res.raise_for_status()
        body = res.content
        if len(body) > PDF_MAX_MB * 1024 * 1024:
            return None
        ctype = (res.headers.get("Content-Type") or "").lower()
        if not body.startswith(b"%PDF") and "pdf" not in ctype:
            return None
        return body
    except requests.RequestException:
        return None


def pdf_text(body: bytes) -> str:
    """앞쪽 페이지 텍스트를 반환. 읽지 못하면 빈 문자열.

    스캔본이라 텍스트가 아예 없는 공고문도 있다. 그때는 공고명 점수로만 판정된다.
    """
    if not HAS_PDF:
        return ""
    try:
        reader = PdfReader(io.BytesIO(body))
        return "\n".join((p.extract_text() or "")
                         for p in reader.pages[:PDF_MAX_PAGES])
    except Exception:  # noqa: BLE001  파싱 실패는 모두 무시
        return ""


def qualification_section(text: str) -> str:
    """입찰참가자격으로 보이는 구간을 전부 모아 돌려준다. 못 찾으면 전체.

    첫 매치만 쓰면 안 된다. 공고문 앞쪽 안내문의
    '입찰참가자격 등록, 나라장터시스템: 정부조달콜센터...' 에 먼저 걸리고
    바로 뒤에 QUAL_TAIL('제출서류')이 나오면 53자만 잘려 나와,
    정작 뒤에 있는 진짜 자격요건을 통째로 못 본다. 검증용 8건 중 2건이 그랬다.

    구간을 좁히는 것 자체는 유지한다. '3865' 처럼 짧은 숫자가 전화번호나
    금액에서 우연히 걸리는 것을 막아주기 때문이다.
    """
    parts = []
    for m in QUAL_HEAD.finditer(text):
        rest = text[m.end():]
        t = QUAL_TAIL.search(rest)
        parts.append(rest[:t.start()] if t else rest[:3000])
    # 구분자는 norm() 이 지우지 않는 문자여야 한다.
    # 공백으로 이으면 두 구간이 붙어 없던 키워드가 생길 수 있다.
    return "|".join(parts) if parts else text


def scan_pdf(raw: dict, sc: Scorer) -> tuple[list[str], bytes | None]:
    """첨부 PDF 에서 자격요건 키워드를 찾는다.

    (찾은 키워드, 저장용 PDF 본문) 반환.
    저장용은 키워드가 나온 그 문서이고, 아무 데서도 안 나오면 첫 번째 PDF 다.
    보통 첫 번째가 입찰공고문이라 사람이 열어볼 것도 그것이다.
    """
    first: bytes | None = None
    for i in range(1, 11):
        url = (raw.get(f"ntceSpecDocUrl{i}") or "").strip()
        if not url:
            continue
        name = (raw.get(f"ntceSpecFileNm{i}") or "").lower()
        if name and not name.endswith(".pdf"):
            continue  # HWP 등은 pypdf 로 못 읽는다
        body = fetch_pdf(url)
        if not body:
            continue
        if first is None:
            first = body
        found = sc.pdf_hit(qualification_section(pdf_text(body)))
        if found:
            return found, body
    return [], first


# ---------------------------------------------------------------------------
# 공고문 PDF 저장 — 엑셀에 오른 건만, 한 건당 한 개
# ---------------------------------------------------------------------------


def safe_name(text: str, limit: int = 60) -> str:
    cleaned = BAD_CHARS.sub("_", str(text)).strip().strip(".")
    cleaned = re.sub(r"\s+", " ", cleaned)
    if len(cleaned) > limit:
        cleaned = cleaned[:limit].rstrip()
    return cleaned or "무제"


def fit_filename(folder: Path, fname: str) -> str:
    """폴더 경로가 길면 확장자를 지키면서 파일명만 줄인다."""
    room = MAX_PATH - len(str(folder)) - 1
    if len(fname) <= room:
        return fname
    stem, dot, ext = fname.rpartition(".")
    if dot and len(ext) <= 8:
        keep = max(6, room - len(ext) - 1)
        return f"{stem[:keep]}.{ext}"
    return fname[:max(8, room)]


def save_pdf(body: bytes, org: str, title: str, no: str, root: Path) -> Path | None:
    """공고문 PDF 를 `공고문/기관_공고명_공고번호.pdf` 로 저장하고 그 경로를 준다.

    공고번호를 붙이는 이유: 같은 기관이 같은 이름으로 재공고하는 일이 흔해서
    기관_공고명 만으로는 앞선 건을 덮어쓴다.
    이미 있으면 None (메일 첨부는 이번에 새로 받은 것만 붙인다).
    """
    folder = root / PDF_DIR
    try:
        folder.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        log(f"    [실패] 공고문 폴더를 만들 수 없습니다 — {exc}")
        return None

    stem = f"{safe_name(org, 30)}_{safe_name(title, 70)}_{safe_name(no, 20)}"
    dest = folder / fit_filename(folder, f"{stem}.pdf")
    if dest.exists() and dest.stat().st_size > 0:
        return None
    try:
        tmp = dest.with_suffix(".pdf.part")
        tmp.write_bytes(body)
        tmp.replace(dest)
        return dest
    except OSError as exc:
        log(f"    [실패] 공고문을 저장하지 못했습니다 — {exc}")
        return None


# ---------------------------------------------------------------------------
# 실행 잠금 · 수집 이력
# ---------------------------------------------------------------------------


class RunLock:
    """공유폴더를 여러 대가 동시에 건드려 엑셀/이력이 깨지는 것을 막는다."""

    STALE_HOURS = 3

    def __init__(self, folder: Path):
        self.path = folder / "_실행중.lock"

    def __enter__(self):
        if self.path.exists():
            age = time.time() - self.path.stat().st_mtime
            if age < self.STALE_HOURS * 3600:
                owner = self.path.read_text(encoding="utf-8", errors="replace").strip()
                sys.exit(
                    f"다른 PC에서 이미 실행 중입니다.\n  {owner}\n"
                    f"  끝난 뒤 다시 실행하세요.\n"
                    f"  (비정상 종료로 남은 파일이면 삭제: {self.path})"
                )
            self.path.unlink(missing_ok=True)

        import socket
        stamp = f"{socket.gethostname()} / {datetime.now(KST):%Y-%m-%d %H:%M}"
        try:
            self.path.write_text(stamp, encoding="utf-8")
        except OSError:
            pass   # 잠금에 실패해도 수집 자체는 진행
        return self

    def __exit__(self, *exc):
        self.path.unlink(missing_ok=True)
        return False


def 이력키(출처: str, no: str) -> str:
    """수집이력의 열쇠. 공고번호는 사이트마다 형식이 달라 출처를 앞에 붙인다."""
    return f"{출처}:{no}"


def load_seen(path: Path) -> dict:
    """수집이력을 {이력키: 지난번 값} 으로 읽는다.

    번호만 남기던 때는 '이미 보낸 공고인가' 만 답할 수 있었다. 그래서 이미
    보낸 공고의 마감이 미뤄져도 알 길이 없었다. 무엇이 바뀌었는지 말하려면
    지난번 값을 들고 있어야 한다.

    옛 파일도 그대로 읽는다. 값이 없는 항목은 '비교 불가' 로 두고, 다음
    회차에 값을 채운다. 넘어온 직후 회차에서 전부 '변경' 으로 쏟아지는 것을
    막기 위해서다 (그건 변경이 아니라 우리가 몰랐던 것뿐이다).
    """
    if not path.exists():
        return {}
    try:
        saved = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}

    if isinstance(saved, dict) and isinstance(saved.get("공고"), dict):
        return {str(k): dict(v) for k, v in saved["공고"].items()
                if isinstance(v, dict)}

    # 옛 형식: ["공고번호", ...] 또는 ["공고번호-000", ...]
    옛 = {}
    for x in saved if isinstance(saved, list) else []:
        no = str(x).rpartition("-")[0] or str(x)
        옛[이력키(SRC_G2B, no)] = {}      # 값이 없다 = 비교할 수 없다
    return 옛


def save_seen(path: Path, seen: dict) -> None:
    try:
        path.write_text(
            json.dumps({"형식": 2, "공고": dict(sorted(seen.items()))},
                       ensure_ascii=False, indent=1),
            encoding="utf-8")
    except OSError as exc:
        log(f"[주의] 수집이력을 저장하지 못했습니다: {exc}")


def first_snapshot(raw: dict) -> dict:
    """옛 이력에서 넘어온 항목을 지금 값으로 채운다.

    등급과 점수는 비워둔다. 그때 어떤 판정이었는지 모르기 때문이다.
    이 항목은 이번 회차에 '변경' 으로 알리지 않는다. 값이 달라서가 아니라
    우리가 몰랐던 것뿐이라 알릴 내용이 없다.
    """
    return {"차수": notice_ord(raw),
            "마감": deadline(raw),
            "예산": money(raw.get("asignBdgtAmt")),
            "공고명": (raw.get("bidNtceNm") or "").strip(),
            "정확도": "", "점수": 0, "본때": RUN_STAMP}


def snapshot(row: list, 차수: int) -> dict:
    """다음 회차에 견줄 값. 엑셀 한 줄에서 필요한 것만 뽑는다.

    차수는 엑셀에 없어서(사람이 볼 값이 아니다) 따로 받는다.
    """
    return {"차수": int(차수),
            "마감": str(row[COL_DUE] or ""),
            "예산": row[COL_BUDGET] if isinstance(row[COL_BUDGET], int) else "",
            "공고명": str(row[COL_TITLE] or ""),
            "정확도": str(row[COL_GRADE] or ""),
            "점수": row[COL_SCORE] if isinstance(row[COL_SCORE], int) else 0,
            "본때": str(row[COL_STAMP] or "")}


# ---------------------------------------------------------------------------
# 엑셀 (실행할 때마다 아래에 줄 추가)
# ---------------------------------------------------------------------------

# 엑셀은 글꼴 후보를 못 쓴다. 이름 하나만 적을 수 있어서, 없는 PC 에서는
# 엑셀이 알아서 비슷한 글꼴로 대신 그린다. 윈도우에 등록된 이름 그대로 적는다.
XLS_FONT = "NICE 고딕Neo2유니 TTF 03 Rg"
XLS_BOLD = "NICE 고딕Neo2유니 TTF 05 Sb"   # 굵은 칸(머리글·등급)에 쓴다
XLS_SIZE = 10


def apply_font(wb) -> None:
    """글꼴을 따로 안 준 칸(=본문 대부분)의 기본 글꼴을 바꾼다.

    통합문서가 들고 있는 글꼴 목록의 0번이 그 기본값이다. 이걸 갈아끼우면
    이미 쌓인 줄까지 한 번에 바뀐다. openpyxl 속내를 건드리는 방법이라,
    안 되면 저장은 그대로 하고 글꼴만 포기한다 (엑셀이 안 나가면 안 된다).
    """
    base = Font(name=XLS_FONT, size=XLS_SIZE)
    try:
        fonts = wb._fonts                        # noqa: SLF001  공개 통로가 없다
        fonts[0] = base
        # 0번을 바꾸면 '글꼴→번호' 색인표가 어긋난다. 다시 만들어 둔다.
        fonts._dict = {}
        for i, f in enumerate(fonts):
            fonts._dict.setdefault(f, i)
        wb._named_styles["Normal"].font = base   # 엑셀의 '표준' 서식도 맞춘다
    except (AttributeError, IndexError, KeyError, TypeError):
        pass


def refit_fonts(ws) -> None:
    """예전 파일에 남아 있는 줄의 글꼴만 새 글꼴로 맞춘다.

    색·밑줄(등급 색, 링크 밑줄)은 그대로 둔다. 굵던 칸은 05 Sb 로 바꾸고
    굵기 표시는 뗀다 — 세미볼드에 굵기를 또 걸면 두 번 굵어진다.
    이걸 안 하면 한 파일 안에서 예전 줄과 새 줄의 서체가 갈린다.
    """
    for row in ws.iter_rows():
        for cell in row:
            f = cell.font
            if f is None or f.name in (XLS_FONT, XLS_BOLD):
                continue                          # 이미 바꾼 칸
            굵음 = bool(f.bold)
            cell.font = Font(name=XLS_BOLD if 굵음 else XLS_FONT,
                             size=XLS_SIZE, bold=False, italic=f.italic,
                             underline=f.underline, color=f.color)


def open_sheet(path: Path):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        old = [c.value for c in ws[1]][:len(HEADERS)]
        if old == HEADERS:
            apply_font(wb)
            refit_fonts(ws)
            return wb, ws
        # 컬럼 구성이 바뀐 옛 파일. 그대로 이어붙이면 값이 한 칸씩 밀린다.
        wb.close()
        backup = path.with_name(f"{path.stem}_이전_{RUN_STAMP[:10]}{path.suffix}")
        try:
            path.replace(backup)
            log(f"[안내] 컬럼 구성이 바뀌어 이전 파일을 보관했습니다: {backup.name}")
        except OSError as exc:
            log(f"[주의] 이전 엑셀을 옮기지 못했습니다: {exc}")

    wb = Workbook()
    apply_font(wb)
    ws = wb.active
    ws.title = "공고목록"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(name=XLS_BOLD, size=XLS_SIZE, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return wb, ws


def append_rows(path: Path, rows: list[list]) -> bool:
    if not rows:
        return True
    wb, ws = open_sheet(path)
    start = ws.max_row + 1
    for row in rows:
        ws.append(row)
    for r in ws.iter_rows(min_row=start):
        g = r[0].value
        r[0].font = Font(name=XLS_BOLD, size=XLS_SIZE,
                         color=GRADE_COLOR.get(g, "000000"))
        r[0].alignment = Alignment(horizontal="center")
        # 검토 필요(C)는 수집분과 같은 파일에 섞이므로 행 전체에 색을 깐다.
        fill = GRADE_FILL.get(g)
        if fill:
            for cell in r:
                cell.fill = PatternFill("solid", fgColor=fill)
        for cell in (r[COL_BUDGET], r[COL_PRICE]):
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        if r[COL_LINK].value:
            r[COL_LINK].hyperlink = r[COL_LINK].value
            r[COL_LINK].value = "공고 열기"
            r[COL_LINK].font = Font(name=XLS_FONT, size=XLS_SIZE,
                                    color="0563C1", underline="single")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    try:
        wb.save(path)
        return True
    except PermissionError:
        log(f"\n[주의] 엑셀이 열려 있어 저장하지 못했습니다: {path}")
        log("       파일을 닫고 다시 실행하세요. 첨부파일은 이미 받아뒀습니다.")
        return False


# ---------------------------------------------------------------------------
# 메일 공지 — 본문에 표를 그대로 넣는다 (요구사항_v3 회신 2번)
# ---------------------------------------------------------------------------

# 사내 서체. 받는 사람 PC 에 이 글꼴이 깔려 있어야 그대로 보이고,
# 없으면 뒤의 맑은 고딕으로 자동으로 넘어간다.
# 윈도우에 등록된 이름은 '고딕' 과 'Neo2유니' 사이가 붙어 있다(오타 아님).
# 영문 이름(GtNeo2Uni)도 함께 적는다. 같은 글꼴을 영문으로 잡는 PC 가 있다.
FONT = ("'NICE 고딕Neo2유니 TTF 03 Rg','NICE GtNeo2Uni TTF 03 Rg',"
        "'맑은 고딕','Malgun Gothic',sans-serif")

# 굵은 글씨는 03 Rg 를 굵게 흉내 내지 않고 05 Sb(세미볼드) 자체를 쓴다.
# 05 Sb 에 font-weight:bold 를 또 걸면 두 번 굵어지므로, 이 글꼴을 쓰는
# 자리에는 font-weight:normal 을 같이 적는다 (아래 BOLD 상수).
FONT_SB = ("'NICE 고딕Neo2유니 TTF 05 Sb','NICE GtNeo2Uni TTF 05 Sb',"
           "'맑은 고딕','Malgun Gothic',sans-serif")
BOLD = f"font-family:{FONT_SB};font-weight:normal"
MAIL_HEADERS = ["구분", "정확도", "공고종류", "등록일", "수요기관", "공고명",
                "마감일시", "사업예산", "검색 키워드"]
# 변경·취소는 볼 것이 다르다. 점수나 키워드가 아니라 '무엇이 바뀌었나' 다.
CHANGE_HEADERS = ["구분", "공고종류", "수요기관", "공고명", "변경내역", "마감일시"]

# 메일의 '사업예산' 은 배정예산금액(asignBdgtAmt)입니다. 부가세를 포함한,
# 발주기관이 잡아둔 돈입니다. 입찰 기준이 되는 추정가격(부가세 제외)은
# 표가 넓어져 메일에는 넣지 않고 엑셀에만 둡니다.

# 점수 근거 중 '어떤 낱말에 걸렸는지' 를 알려주는 항목들.
# '조합+4' '기관+3' 같은 계산 내역은 메일에 넣지 않는다.
WHY_KEYWORD_TAGS = ("확정: ", "채권시장: ", "대상: ", "행위: ")


def search_keywords(row: list) -> str:
    """메일의 '검색 키워드' 칸.

    A 는 업종코드/채권평가회사가 그대로 들어간다. B·C 는 키워드 칸이 비어
    있으므로 점수 근거에서 걸린 낱말만 뽑아 쓴다. 배점 계산 내역은 뺀다.
    """
    kw = str(row[COL_KW] or "").strip()
    if kw and kw != "-":
        return kw
    found: list[str] = []
    for part in str(row[COL_WHY] or "").split(" · "):
        for tag in WHY_KEYWORD_TAGS:
            if part.startswith(tag):
                for word in part[len(tag):].split(", "):
                    if word and word not in found:
                        found.append(word)
    # 점수표에 '공정가치평가' '공정가치' '평가' 가 다 들어 있어 한 제목에서
    # 셋이 같이 걸린다. 짧은 쪽은 군더더기라 긴 낱말만 남긴다.
    keep = [w for w in found if not any(w != o and w in o for o in found)]
    return ", ".join(keep[:4])


def esc(text) -> str:
    """메일 본문에 넣기 전 HTML 특수문자를 막는다.

    공고명에 '&' 나 '<' 가 들어가는 일이 실제로 있다. 그대로 넣으면
    표가 깨지거나 뒷부분이 사라진다.
    """
    return (str(text or "")
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;"))


def won(value) -> str:
    """메일 표에 넣을 금액. 천 단위마다 쉼표, 없으면 '-'."""
    try:
        return f"{int(value):,}"
    except (TypeError, ValueError):
        return "-"


def half_of_day(when: datetime) -> str:
    """오전/오후 두 번 공지하므로 제목으로 구분이 돼야 한다."""
    return "오전" if when.hour < 12 else "오후"


def day_stamp(when: datetime) -> str:
    """메일에 쓰는 날짜 표기. `08.13(오전)` 로 통일한다."""
    return f"{when:%m.%d}({half_of_day(when)})"


def fill(text: dict, key: str, **vals) -> str:
    """config.yaml 의 문구에 값을 채운다.

    직원이 고치는 값이라 자리표시자를 잘못 적을 수 있다. 그때 예외로 죽으면
    메일이 통째로 안 나가므로, 그 줄만 기본 문구로 되돌리고 알려만 준다.
    """
    tpl = text.get(key) or MAIL_TEXT[key]
    for candidate in (tpl, MAIL_TEXT[key]):
        try:
            return str(candidate).format(**vals)
        except (KeyError, IndexError, ValueError) as exc:
            log(f"[주의] mail.문구.{key} 를 쓸 수 없어 기본 문구로 보냅니다: {exc}")
            log(f"       쓸 수 있는 자리표시자: "
                f"{', '.join('{' + k + '}' for k in vals)}")
    return MAIL_TEXT[key]


def split_rows(rows: list[list]) -> tuple[list[list], list[list]]:
    """(새로 올라온 것, 이미 보낸 것 중 바뀐 것).

    가르는 기준은 공고종류가 아니라 변경내역이 채워졌는지다. 처음 보는
    공고인데 나라장터가 '변경공고' 라고 달아둔 경우가 있는데, 그건 우리에겐
    새 공고다. 우리 이력과 견줘 달라진 것만 변경으로 본다.
    """
    바뀜 = [r for r in rows if str(r[COL_DIFF] or "").strip()]
    새것 = [r for r in rows if not str(r[COL_DIFF] or "").strip()]
    return 새것, 바뀜


def summary_parts(rows: list[list]) -> tuple[int, str]:
    """(건수, '수집 9건 · 검토 4건 · 변경 2건')"""
    새것, 바뀜 = split_rows(rows)
    n_c = sum(1 for r in 새것 if r[COL_GRADE] == "C")
    n_hit = len(새것) - n_c
    n_취소 = sum(1 for r in 바뀜 if r[COL_KIND] == "취소")
    n_변경 = len(바뀜) - n_취소
    내역 = " · ".join(x for x in (f"수집 {n_hit}건" if n_hit else "",
                                 f"검토 {n_c}건" if n_c else "",
                                 f"변경 {n_변경}건" if n_변경 else "",
                                 f"취소 {n_취소}건" if n_취소 else "") if x)
    return len(rows), 내역


def no_news_line(text: dict, when: datetime) -> str:
    """0건일 때 쓰는 문구. 이 한 줄이 메일 내용 전부다."""
    return fill(text, "없음", 날짜=day_stamp(when))


def mail_subject(cfg: dict, rows: list[list], when: datetime,
                 status: str = 정상) -> str:
    text = cfg["mail"]["문구"]
    stamp = day_stamp(when)
    건수, 내역 = summary_parts(rows)
    if status == 일부실패:
        # 0건이어도 '없습니다' 라고 하면 안 된다. 못 본 것일 수 있다.
        return fill(text, "제목_일부", 날짜=stamp, 건수=건수,
                    내역=내역 or "수집 0건")
    if not rows:
        return fill(text, "제목_없음", 날짜=stamp)
    if split_rows(rows)[1]:
        # 변경·취소가 섞였으면 전부를 '신규' 라고 부를 수 없다.
        return fill(text, "제목_변경", 날짜=stamp, 건수=건수, 내역=내역)
    return fill(text, "제목", 날짜=stamp, 건수=건수, 내역=내역)


def period_phrase(begin: datetime, end: datetime) -> str:
    """이번 실행이 훑은 기간. 수집을 직접 했을 때만 쓸 수 있다.

    시각까지 적는다. 오전/오후 두 번 돌리면 날짜만으로는 두 메일이
    같은 기간을 본 것처럼 보인다. 같은 날이면 뒤쪽 날짜는 생략한다.
    """
    if begin.date() == end.date():
        return f"조회기간 {begin:%m.%d %H:%M} ~ {end:%H:%M}"
    return f"조회기간 {begin:%m.%d %H:%M} ~ {end:%m.%d %H:%M}"


def reg_range_phrase(rows: list[list]) -> str:
    """표에 오른 공고의 등록일 범위.

    --메일만 은 수집을 하지 않으므로 그 회차가 어느 기간을 훑었는지 알 수 없다.
    (엑셀에 조회기간을 남기지 않는다.) config 의 hours 로 거꾸로 계산하면
    지금 설정값일 뿐이어서 실제와 어긋난다. 그래서 자료에 있는 사실만 쓴다.
    """
    days = sorted(str(r[COL_REG])[:10] for r in rows if r[COL_REG])
    if not days:
        return ""
    def md(s: str) -> str:
        return s[5:].replace("-", ".")
    return (f"공고 등록일 {md(days[0])} ~ {md(days[-1])}"
            if days[0] != days[-1] else f"공고 등록일 {md(days[0])}")


def as_html(line: str) -> str:
    """문구 한 덩어리를 본문에 넣을 수 있게 바꾼다. 줄바꿈은 <br> 로."""
    return esc(line).replace("\n", "<br>")


def change_table(rows: list[list], th: str, td: str, text: dict) -> list[str]:
    """변경·취소 표. 신규 표보다 먼저 나와야 한다.

    마감이 미뤄지거나 공고가 취소된 것은 이미 검토를 시작한 건이라
    새 공고보다 급하다. 아래에 붙이면 표를 다 읽은 뒤에야 보게 된다.
    """
    if not rows:
        return []
    parts = [f'<p style="margin:14px 0 6px;{BOLD};color:#c00000">'
             f'{as_html(fill(text, "변경머리", 건수=len(rows)))}</p>',
             '<table cellspacing="0" cellpadding="0" '
             f'style="border-collapse:collapse;font-size:10pt;'
             f'font-family:{FONT}">',
             "<tr>" + "".join(f'<th style="{th}">{esc(h)}</th>'
                              for h in CHANGE_HEADERS) + "</tr>"]
    for no, r in enumerate(rows, start=1):
        kind = str(r[COL_KIND] or "-")
        color = KIND_COLOR.get(kind, "#888")
        취소 = kind == "취소"
        bg = "background:#fdf0f0;" if 취소 else ""
        title = esc(r[COL_TITLE])
        if r[COL_LINK]:
            title = (f'<a href="{esc(r[COL_LINK])}" style="color:#0563c1">'
                     f'{title}</a>')
        마감 = "—" if 취소 else (esc(r[COL_DUE]) or "-")
        parts.append(
            f'<tr style="{bg}">'
            f'<td style="{td};text-align:center;color:#888">{no}</td>'
            f'<td style="{td};text-align:center;white-space:nowrap;'
            f'color:{color};{BOLD}">{esc(kind)}</td>'
            f'<td style="{td}">{esc(r[COL_DM] or r[COL_NT] or "기관미상")}</td>'
            f'<td style="{td};max-width:360px">{title}</td>'
            f'<td style="{td};color:#c00000;max-width:300px">'
            f'{esc(r[COL_DIFF])}</td>'
            f'<td style="{td};white-space:nowrap">{마감}</td>'
            "</tr>")
    parts.append("</table>")
    return parts


def new_table(rows: list[list], th: str, td: str) -> list[str]:
    """새로 올라온 공고 표. 이 메일의 본체다."""
    # 아웃룩은 표 안에서 바깥 글꼴을 물려받지 않는다. 표에도 다시 적는다.
    parts = ['<table cellspacing="0" cellpadding="0" '
             f'style="border-collapse:collapse;font-size:10pt;'
             f'font-family:{FONT}">',
             "<tr>" + "".join(f'<th style="{th}">{esc(h)}</th>'
                              for h in MAIL_HEADERS) + "</tr>"]
    for no, r in enumerate(rows, start=1):
        g = r[COL_GRADE]
        bg = "background:#f5f5f5;" if g == "C" else ""
        color = "#" + GRADE_COLOR.get(g, "000000")
        org = r[COL_DM] or r[COL_NT] or "기관미상"
        title = esc(r[COL_TITLE])
        kind = str(r[COL_KIND] or "-")
        kind_color = KIND_COLOR.get(kind, "#888")
        if r[COL_LINK]:
            title = (f'<a href="{esc(r[COL_LINK])}" style="color:#0563c1">'
                     f'{title}</a>')
        parts.append(
            f'<tr style="{bg}">'
            f'<td style="{td};text-align:center;color:#888">{no}</td>'
            f'<td style="{td};text-align:center;white-space:nowrap;'
            f'color:{color};{BOLD}">{esc(GRADE_LABEL.get(g, g))}</td>'
            f'<td style="{td};text-align:center;white-space:nowrap;'
            f'color:{kind_color}">{esc(kind)}</td>'
            f'<td style="{td};white-space:nowrap">{esc(r[COL_REG])}</td>'
            f'<td style="{td}">{esc(org)}</td>'
            f'<td style="{td};max-width:420px">{title}</td>'
            f'<td style="{td};white-space:nowrap">{esc(r[COL_DUE]) or "-"}</td>'
            f'<td style="{td};text-align:right;white-space:nowrap">'
            f'{won(r[COL_BUDGET])}</td>'
            f'<td style="{td};color:#666;max-width:300px">'
            f'{esc(search_keywords(r))}</td>'
            "</tr>")
    parts.append("</table>")
    return parts


def mail_html(rows: list[list], period: str, end: datetime, root: Path,
              attached: list[Path], text: dict, 빠짐: str = "") -> str:
    """본문 HTML. 아웃룩이 지원하는 범위(표 + 인라인 스타일)만 쓴다.

    빠짐 이 있으면 표보다 먼저 경고를 붙인다. 표를 다 읽고 맨 아래에서야
    '실은 일부만 조회됐습니다' 를 보면 이미 판단이 끝난 뒤다.
    """
    th = ("padding:6px 8px;border:1px solid #d0d0d0;background:#1f3864;"
          f"color:#fff;{BOLD};text-align:center;white-space:nowrap")
    td = "padding:6px 8px;border:1px solid #d0d0d0;vertical-align:top"

    n_c = sum(1 for r in rows
              if r[COL_GRADE] == "C" and not str(r[COL_DIFF] or "").strip())
    # 본문 기본 크기. 첫줄('… 신규 공고입니다')과 요약 줄이 이걸 물려받는다.
    # 아웃룩은 pt 로 적어야 워드에서 보던 크기와 같게 나온다.
    parts = [f'<div style="font-family:{FONT};font-size:11pt;color:#222">']

    if 빠짐:
        경고 = fill(text, "실패알림", 빠짐=빠짐)
        parts.append(
            '<p style="margin:0 0 10px;padding:8px 10px;border:1px solid #c00000;'
            f'background:#fdf0f0;color:#c00000;{BOLD}">'
            f'{as_html(경고)}</p>')

    # 0건이면 한 줄만 보낸다. 조회기간·등급설명은 볼 표가 없으면 군더더기다.
    if not rows and not 빠짐:
        parts.append(f'<p style="margin:0">{as_html(no_news_line(text, end))}</p>')
    elif not rows:
        parts.append('<p style="margin:0">조회된 신규 공고가 없습니다. '
                     '다만 위 구간을 못 받았으므로 없다고 단정할 수 없습니다.</p>')
    else:
        새것, 바뀜 = split_rows(rows)
        건수, 내역 = summary_parts(rows)
        머리 = fill(text, "첫줄", 시각=f"{end:%Y.%m.%d %H:%M}", 기간=period)
        if not period:
            머리 = 머리.replace(" ()", "")   # 기간을 모를 때 빈 괄호가 남는다
        parts.append(f'<p style="margin:0 0 4px">{as_html(머리)}</p>')
        parts.append(f'<p style="margin:0 0 10px;{BOLD}">'
                     f'{as_html(fill(text, "요약", 건수=건수, 내역=내역))}</p>')

        parts += change_table(바뀜, th, td, text)
        if 새것:
            if 바뀜:
                parts.append(f'<p style="margin:16px 0 6px;{BOLD}">'
                             f'{as_html(fill(text, "신규머리", 건수=len(새것)))}</p>')
            parts += new_table(새것, th, td)
        else:
            parts.append('<p style="margin:14px 0 0">새로 올라온 공고는 '
                         '없습니다.</p>')

        parts.append(f'<p style="margin:12px 0 0;font-size:10pt;color:#777">'
                     f'{as_html(fill(text, "안내"))}</p>')
        if n_c:
            parts.append(f'<p style="margin:4px 0 0;font-size:10pt;color:#777">'
                         f'{as_html(fill(text, "검토안내"))}</p>')

    # 0건 메일은 문구 한 줄로 끝낸다. 첨부도 안내문도 붙이지 않는다.
    if rows:
        if attached:
            names = ", ".join(esc(p.name) for p in attached[:6])
            more = f" 외 {len(attached) - 6}개" if len(attached) > 6 else ""
            parts.append(f'<p style="margin:10px 0 0;font-size:10pt;color:#777">'
                         f'첨부: {names}{more}</p>')
        parts.append(f'<p style="margin:10px 0 0;font-size:9.5pt;color:#999">'
                     f'{as_html(fill(text, "꼬리말", 폴더=root))}</p>')
    parts.append("</div>")
    return "\n".join(parts)


def mail_text(rows: list[list], end: datetime, text: dict,
              빠짐: str = "") -> str:
    """HTML 을 못 읽는 메일 앱을 위한 대체 본문."""
    머리 = [fill(text, "실패알림", 빠짐=빠짐), ""] if 빠짐 else []
    if not rows:
        본문 = "조회된 신규 공고가 없습니다." if 빠짐 else no_news_line(text, end)
        return "\n".join(머리 + [본문])
    새것, 바뀜 = split_rows(rows)
    건수, 내역 = summary_parts(rows)
    lines = 머리 + [fill(text, "요약", 건수=건수, 내역=내역), ""]
    if 바뀜:
        lines.append(fill(text, "변경머리", 건수=len(바뀜)))
        for no, r in enumerate(바뀜, start=1):
            lines.append(f"{no}. [{r[COL_KIND]}] "
                         f"{r[COL_DM] or r[COL_NT]} — {r[COL_TITLE]}")
            lines.append(f"    {r[COL_DIFF]} · {r[COL_LINK]}")
        lines += ["", fill(text, "신규머리", 건수=len(새것))]
    for no, r in enumerate(새것, start=1):
        kind = str(r[COL_KIND] or "")
        표시 = f"[{kind}] " if kind in KIND_COLOR else ""   # 신규는 굳이 안 적는다
        lines.append(f"{no}. [{GRADE_LABEL.get(r[COL_GRADE], r[COL_GRADE])}] "
                     f"{표시}{r[COL_DM] or r[COL_NT]} — {r[COL_TITLE]}")
        금액 = won(r[COL_BUDGET])
        예산 = "예산 미기재" if 금액 == "-" else f"예산 {금액}원"
        lines.append(f"    마감 {r[COL_DUE]} · {예산}"
                     f" · {search_keywords(r)} · {r[COL_LINK]}")
    return "\n".join(lines)


def find_saved_pdfs(root: Path, rows: list[list]) -> list[Path]:
    """이 회차 공고들의 공고문 PDF 를 `공고문/` 에서 찾아온다.

    파일명 끝에 공고번호가 붙어 있으므로 그것으로 찾는다. 파일명을 다시
    조립해서 맞춰보는 방법은 기관명·공고명 길이 제한 때문에 어긋나기 쉽다.

    이번 실행에서 새로 받은 것만 쓰지 않는 이유: 같은 공고문이 이미 있으면
    다시 받지 않으므로(save_pdf), 그것만 모으면 첨부가 비는 경우가 생긴다.
    """
    folder = root / PDF_DIR
    if not folder.is_dir():
        return []
    files = sorted(folder.glob("*.pdf"))
    picked: list[Path] = []
    for r in rows:
        no = str(r[COL_NO] or "").strip()
        if not no:
            continue
        for p in files:
            if no in p.name and p not in picked:
                picked.append(p)
                break
    return picked


def pick_attachments(cfg: dict, root: Path, pdfs: list[Path]) -> list[Path]:
    """첨부를 고른다. 용량이 크면 앞에서부터 담고 나머지는 버린다.

    사내 메일은 첨부 용량 제한이 있고, 본문에 표가 들어가므로
    첨부가 잘려도 내용 확인은 된다.
    """
    m = cfg["mail"]
    picked: list[Path] = []
    total = 0.0
    cap = m["attach_max_mb"] * 1024 * 1024

    wanted: list[Path] = []
    if m["attach_excel"]:
        wanted.append(root / COLLECT_XLSX)
    if m["attach_pdf"]:
        wanted.extend(pdfs)

    for p in wanted:
        try:
            size = p.stat().st_size
        except OSError:
            continue
        if total + size > cap:
            log(f"  [안내] 첨부 용량({m['attach_max_mb']:.0f}MB)을 넘어 "
                f"{p.name} 부터는 첨부하지 않습니다.")
            break
        picked.append(p)
        total += size
    return picked


def send_via_outlook(cfg: dict, subject: str, html: str,
                     attachments: list[Path],
                     to: list[str] | None = None,
                     cc: list[str] | None = None) -> bool:
    """설치된 아웃룩으로 보낸다. 사내 메일 서버 정보가 필요 없다.

    로그인한 계정으로 나가므로 발신자가 본인 계정이 된다.
    회사 정책으로 '프로그래밍 방식 액세스'가 막혀 있으면 실패한다.
    """
    try:
        import win32com.client  # type: ignore
    except ImportError:
        # exe 배포판에는 pywin32 가 이미 들어 있다. 여기까지 왔다면 번들이
        # 깨진 것이므로 설치하라고 안내해봐야 도움이 안 된다.
        log("  [실패] 아웃룩 발송에는 pywin32 가 필요합니다:  " +
            ("배포 폴더가 손상됐습니다. 다시 받으세요" if FROZEN
             else "pip install pywin32"))
        return False

    m = cfg["mail"]
    to = to or m["to"]
    cc = m["cc"] if cc is None else cc
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        item = outlook.CreateItem(0)          # 0 = olMailItem
        item.To = "; ".join(to)
        if cc:
            item.CC = "; ".join(cc)
        item.Subject = subject
        item.HTMLBody = html
        for p in attachments:
            item.Attachments.Add(str(p.resolve()))
        if m.get("draft_only"):
            item.Display()                    # 사람이 확인한 뒤 직접 보낸다
            log("  [안내] 아웃룩에 초안을 띄웠습니다. 확인 후 직접 보내세요.")
            return True
        item.Send()
        return True
    except Exception as exc:  # noqa: BLE001  COM 오류 종류가 많다
        log(f"  [실패] 아웃룩 발송 — {type(exc).__name__}: {exc}")
        log("         아웃룩이 실행 중인지, 보안 정책으로 스크립트 발송이")
        log("         막혀 있지 않은지 확인하세요. (mail.mode: smtp 로도 가능)")
        return False


def send_via_smtp(cfg: dict, subject: str, html: str, text: str,
                  attachments: list[Path],
                  to: list[str] | None = None,
                  cc: list[str] | None = None) -> bool:
    """사내 SMTP 릴레이로 보낸다. host·from 은 IT 에서 받아야 한다."""
    import mimetypes
    import smtplib
    from email.message import EmailMessage

    m = cfg["mail"]
    to = to or m["to"]
    cc = m["cc"] if cc is None else cc
    s = m["smtp"]
    host = str(s.get("host") or "").strip()
    sender = str(s.get("from") or s.get("user") or "").strip()
    if not host or not sender:
        log("  [실패] mail.smtp 의 host / from 이 비어 있습니다. IT 에 문의하세요.")
        return False

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = ", ".join(to)
    if cc:
        msg["Cc"] = ", ".join(cc)
    msg.set_content(text)
    msg.add_alternative(html, subtype="html")

    for p in attachments:
        ctype, _ = mimetypes.guess_type(p.name)
        maintype, _, subtype = (ctype or "application/octet-stream").partition("/")
        try:
            msg.add_attachment(p.read_bytes(), maintype=maintype,
                               subtype=subtype, filename=p.name)
        except OSError as exc:
            log(f"  [주의] 첨부를 읽지 못했습니다: {p.name} — {exc}")

    port = int(s.get("port") or 25)
    try:
        with smtplib.SMTP(host, port, timeout=60) as smtp:
            if s.get("tls", port == 587):
                smtp.starttls()
            if s.get("user"):
                smtp.login(str(s["user"]), str(s.get("password") or ""))
            smtp.send_message(msg)
        return True
    except Exception as exc:  # noqa: BLE001  smtplib 예외가 여러 갈래다
        log(f"  [실패] SMTP 발송 — {type(exc).__name__}: {exc}")
        return False


def error_body(checks: list[dict], root: Path, begin: datetime,
               end: datetime) -> tuple[str, str]:
    """오류 메일 본문 (HTML, 글자만). 받는 사람은 담당자가 아니라 관리자다.

    업무 문구가 아니라 진단 정보라서 config 의 문구로 빼지 않았다.
    고칠 사람이 봐야 할 것만 적는다: 무엇이 실패했고, 어느 구간이 빠졌고,
    무엇을 다시 돌리면 되는지.
    """
    줄 = [f"조회기간   {begin:%Y-%m-%d %H:%M} ~ {end:%Y-%m-%d %H:%M}",
         f"실행시각   {RUN_STAMP}",
         f"저장위치   {root}",
         ""]
    for c in checks:
        상태 = "실패" if c["빠짐"] else "정상"
        줄.append(f"[{상태}] {c['이름']}  {c['건수']:,}건")
        for 구간 in c["빠짐"]:
            줄.append(f"         못 받음: {구간}")
    줄 += ["",
          "이 회차에서 빠진 구간은 다음 실행 때 저절로 메워지지 않습니다.",
          "조회기간이 실행할 때마다 앞으로 밀리기 때문입니다. 아래 명령으로",
          "그 회차를 다시 돌려야 그 구간의 공고를 볼 수 있습니다.",
          "",
          f"    {rerun_command(begin, end)}",
          "",
          f"회차별 조회 성공·실패 기록: {root / CHECK_LOG}",
          f"자세한 실행 기록: {LOG_DIR / '실행.log'}"]
    text = "\n".join(줄)
    html = (f'<div style="font-family:{FONT};font-size:11pt;color:#222">'
            f'<p style="margin:0 0 10px;color:#c00000;{BOLD}">'
            f'나라장터 자동수집이 완전하게 끝나지 않았습니다.</p>'
            f'<pre style="margin:0;font-family:Consolas,monospace;'
            f'font-size:10pt;background:#f7f7f7;border:1px solid #ddd;'
            f'padding:10px;white-space:pre-wrap">{esc(text)}</pre></div>')
    return html, text


def notify_error(cfg: dict, checks: list[dict], root: Path,
                 begin: datetime, end: datetime) -> None:
    """조회가 통째로 실패한 회차를 관리자에게 알린다.

    결과 메일은 보내지 않는다. 0건짜리 결과 메일이 나가면 받는 사람이
    '오늘은 공고가 없었구나' 로 읽어버리기 때문이다. 그게 이 기능의 전부다.
    """
    m = cfg["mail"]
    if not m["enabled"]:
        return
    받는이 = m["error_to"] or m["to"]
    if not 받는이:
        return
    subject = fill(m["문구"], "제목_실패", 날짜=day_stamp(end))
    html, text = error_body(checks, root, begin, end)

    log(f"\n[오류메일] {m['mode']} · 수신 {', '.join(받는이)}")
    if m["mode"] == "smtp":
        ok = send_via_smtp(cfg, subject, html, text, [], to=받는이, cc=[])
    else:
        ok = send_via_outlook(cfg, subject, html, [], to=받는이, cc=[])
    log(f"  {'보냈습니다' if ok else '보내지 못했습니다'}: {subject}")
    if not ok:
        # 메일 자체가 막혀 있으면 오류 메일도 못 나간다. 그때 마지막으로
        # 남는 것이 로그와 조회기록 파일이다.
        log("  [주의] 오류 메일도 나가지 못했습니다."
            f" {root / CHECK_LOG} 를 확인하세요.")


def notify(cfg: dict, rows: list[list], root: Path,
           end: datetime, period: str, 빠짐: str = "",
           status: str = 정상) -> None:
    """수집 결과를 메일로 보낸다. 실패해도 수집 결과는 이미 저장돼 있다."""
    m = cfg["mail"]
    if not m["enabled"]:
        return
    if not rows and not m["send_when_empty"] and not 빠짐:
        log("\n[메일] 신규 공고가 없어 보내지 않았습니다."
            " (mail.send_when_empty 로 바꿀 수 있습니다)")
        return

    # 첨부할 공고문은 표에 오른 행에서 되찾는다. 수집 직후든 --메일만 이든
    # 같은 방식으로 찾아야 한쪽만 첨부가 비는 일이 없다.
    # 0건이면 '특이사항 없음' 한 줄만 보내므로 첨부도 붙이지 않는다.
    attachments: list[Path] = []
    if rows:
        pdfs = find_saved_pdfs(root, rows) if m["attach_pdf"] else []
        attachments = pick_attachments(cfg, root, pdfs)
    subject = mail_subject(cfg, rows, end, status)
    html = mail_html(rows, period, end, root, attachments, m["문구"], 빠짐)
    text = mail_text(rows, end, m["문구"], 빠짐)

    log(f"\n[메일] {m['mode']} · 수신 {', '.join(m['to'])}"
        + (f" · 참조 {', '.join(m['cc'])}" if m["cc"] else "")
        + (f" · 첨부 {len(attachments)}개" if attachments else ""))
    if m["mode"] == "smtp":
        ok = send_via_smtp(cfg, subject, html, text, attachments)
    else:
        ok = send_via_outlook(cfg, subject, html, attachments)
    log(f"  {'보냈습니다' if ok else '보내지 못했습니다'}: {subject}")


# ---------------------------------------------------------------------------
# 메일만 다시 보내기 (--메일만)
# ---------------------------------------------------------------------------


def column_map(head: list[str]) -> dict[int, int] | None:
    """엑셀 첫 줄의 이름을 보고 {지금 자리: 그 파일에서의 자리} 를 만든다.

    컬럼을 늘릴 때마다 예전 파일을 못 읽게 되면 --메일만 으로 재발송을 할 수
    없다. 자리를 세는 대신 이름을 맞춰보면 칸이 늘어나도 그대로 읽힌다.
    쓸모없는 파일(다른 엑셀을 잘못 지정한 경우)이면 None 을 준다.
    """
    자리 = {}
    for pos, name in enumerate(head):
        name = RENAMED.get(name, name)
        if name in HEADERS:
            자리.setdefault(HEADERS.index(name), pos)
    # 이 넷이 없으면 메일 표를 못 그린다. 우리 엑셀이 아니라고 본다.
    필수 = (COL_GRADE, COL_TITLE, COL_LINK, COL_STAMP)
    return None if any(c not in 자리 for c in 필수) else 자리


def read_saved_rows(path: Path) -> list[list]:
    """이미 저장된 엑셀에서 행을 그대로 읽어온다.

    링크 칸은 '공고 열기' 라는 글자로 바뀌어 있고 실제 주소는 하이퍼링크에
    들어 있으므로, 그 주소를 다시 꺼내야 메일에서 링크가 살아난다.
    """
    if not path.exists():
        return []
    try:
        wb = load_workbook(path)
    except Exception as exc:  # noqa: BLE001  손상된 파일 등
        log(f"[주의] 엑셀을 읽지 못했습니다: {path.name} — {exc}")
        return []

    ws = wb.active
    head = [str(c.value or "").strip() for c in ws[1]]
    자리 = column_map(head)
    if 자리 is None:
        log(f"[주의] 컬럼 구성이 달라 건너뜁니다: {path.name}")
        return []
    빠진칸 = [h for i, h in enumerate(HEADERS) if i not in 자리]
    if 빠진칸:
        log(f"[안내] 컬럼이 늘기 전 파일입니다. {', '.join(빠진칸)} 은 빈칸으로"
            f" 읽습니다: {path.name}")

    rows = []
    for r in ws.iter_rows(min_row=2):
        raw = [c.value for c in r]
        vals = [""] * len(HEADERS)
        for 새, 옛 in 자리.items():
            if 옛 < len(raw):
                vals[새] = raw[옛]
        if not vals[COL_GRADE]:
            continue
        vals[COL_GRADE] = OLD_GRADE.get(vals[COL_GRADE], vals[COL_GRADE])
        cell = r[자리[COL_LINK]]
        vals[COL_LINK] = (cell.hyperlink.target if cell.hyperlink
                          else cell.value) or ""
        rows.append(vals)
    return rows


def mail_only(cfg: dict, root: Path) -> int:
    """수집은 건너뛰고 엑셀의 마지막 회차를 메일로 보낸다.

    발송 방식을 시험하거나 메일이 실패해 다시 보낼 때 쓴다. API 호출 한도가
    있으므로 그때마다 수집을 처음부터 돌릴 수는 없다.
    """
    # 예전에 파일이 두 개로 갈렸던 시절 자료도 함께 읽어 한 표로 합친다.
    rows = read_saved_rows(root / COLLECT_XLSX) + read_saved_rows(root / OLD_REVIEW_XLSX)
    if not rows:
        log(f"보낼 내용이 없습니다. 먼저 수집을 한 번 돌리세요.\n  {root / COLLECT_XLSX}")
        return 1

    stamps = sorted({str(r[COL_STAMP]) for r in rows if r[COL_STAMP]})
    latest = stamps[-1] if stamps else ""
    batch = [r for r in rows if str(r[COL_STAMP]) == latest] or rows
    batch.sort(key=lambda r: (GRADE_ORDER.get(r[COL_GRADE], 9),
                              -int(r[COL_SCORE] or 0)))

    try:
        end = datetime.strptime(latest, "%Y-%m-%d %H:%M").replace(tzinfo=KST)
    except ValueError:
        end = datetime.now(KST)

    log(f"[메일만] 수집을 건너뛰고 {latest or '마지막'} 회차 "
        f"{len(batch)}건을 보냅니다.")
    notify(cfg, batch, root, end, reg_range_phrase(batch))
    return 0


# ---------------------------------------------------------------------------
# 판정 — 요구사항 1~6차
# ---------------------------------------------------------------------------


def judge(raw: dict, cfg: dict, sc: Scorer, lic_index: dict,
          use_pdf: bool) -> tuple[str, int, str, str, bytes | None] | None:
    """(등급, 점수, 키워드, 근거, 저장용 PDF) 또는 None."""
    uid = (raw.get("bidNtceNo") or "").strip()
    title = (raw.get("bidNtceNm") or "").strip()
    orgs = f"{raw.get('dminsttNm') or ''} {raw.get('ntceInsttNm') or ''}"

    pts, hits = sc.score(title, orgs)
    why = " · ".join(hits)

    # 1차 — 면허제한에 지정 업종코드가 있으면 점수와 무관하게 무조건 수집.
    # 판정에는 PDF 가 필요없지만 사람이 열어볼 공고문은 남겨야 하므로 받아둔다.
    codes, labels = lic_index.get(uid, (set(), []))
    hit_codes = codes & set(cfg["industry_codes"])
    if hit_codes:
        kw = "업종코드 " + ", ".join(sorted(hit_codes))
        # 걸린 업종을 앞에 놓는다. 허용업종이 여러 개일 때 뒤로 밀려
        # 잘려나가면 왜 수집됐는지 안 보인다.
        맞음 = [x for x in labels if any(c in x for c in hit_codes)]
        나머지 = [x for x in labels if x not in 맞음]
        note = "면허제한"
        if labels:
            note += f"({len(labels)}개 허용): " + ", ".join(맞음 + 나머지)[:120]
        _, body = scan_pdf(raw, sc) if use_pdf else ([], None)
        return "A", pts, kw, note, body

    # 2차 — 공고명 점수. 연기금·공제회는 문턱이 더 낮다 (점수.review_cut)
    if pts < sc.review_cut(orgs):
        return None

    # 3·4차 — 첨부 PDF 확인
    body = None
    if use_pdf:
        found, body = scan_pdf(raw, sc)
        if found:
            return "A", pts, ", ".join(found[:3]), "공고문 자격요건 · " + why, body

    # 5차 / 6차
    if pts >= sc.cut_collect:
        return "B", pts, "-", why, body
    return "C", pts, "-", why, body


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def take_option(argv: list[str], name: str) -> tuple[list[str], str | None]:
    """`--이름 값` 과 `--이름=값` 을 모두 받아 (나머지 인자, 값) 을 준다."""
    rest: list[str] = []
    value: str | None = None
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == name and i + 1 < len(argv):
            i += 1
            value = argv[i]
        elif a.startswith(name + "="):
            value = a.split("=", 1)[1]
        else:
            rest.append(a)
        i += 1
    return rest, value


def parse_when(text: str) -> datetime:
    """--기준 에 적은 일시를 읽는다. 날짜만 적으면 자정으로 본다."""
    text = text.strip().strip('"').replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H", "%Y-%m-%d",
                "%Y%m%d%H%M", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=KST)
        except ValueError:
            continue
    sys.exit(f"--기준 을 읽지 못했습니다: {text}\n"
             '  예:  --기준 "2026-08-06 10:00"')


def main() -> int:
    argv = sys.argv[1:]
    argv, 기준_raw = take_option(argv, "--기준")
    argv, 저장_raw = take_option(argv, "--저장")
    args = [a for a in argv if not a.startswith("--")]
    flags = {a for a in argv if a.startswith("--")}
    list_only = "--목록만" in flags
    use_pdf = "--pdf없이" not in flags

    cfg = load_config()
    if 저장_raw:
        # 지난 회차를 재현할 때 진짜 수집이력·엑셀을 건드리지 않게 한다.
        cfg["output_dir"] = resolve_output_dir(저장_raw)
    if "--메일없이" in flags:
        cfg["mail"]["enabled"] = False

    if "--메일만" in flags:
        # 보내려고 실행한 것이므로 mail.enabled 가 false 여도 보낸다.
        if not cfg["mail"]["to"]:
            sys.exit("config.yaml 의 mail.to 에 받는 주소를 먼저 넣으세요.")
        cfg["mail"]["enabled"] = True
        cfg["mail"]["send_when_empty"] = True
        return mail_only(cfg, cfg["output_dir"])

    sc = Scorer()

    hours = cfg["hours"]
    if args:
        try:
            hours = int(float(args[0]) * 24)
        except ValueError:
            예 = ("나라장터수집기.exe 수집 30" if FROZEN
                 else "python 수집기.py 30")
            sys.exit(f"일수는 숫자로 입력하세요. 예: {예}")

    if use_pdf and not HAS_PDF:
        log("[안내] PDF 검사를 하려면 pypdf 가 필요합니다:  " +
            ("배포 폴더가 손상됐습니다. 다시 받으세요" if FROZEN
             else "pip install pypdf"))
        log("       면허제한 + 공고명 점수만으로 진행합니다.\n")
        use_pdf = False

    end = parse_when(기준_raw) if 기준_raw else datetime.now(KST)
    begin = end - timedelta(hours=hours)
    root = cfg["output_dir"]

    if 기준_raw:
        # 엑셀의 수집일시와 메일 제목의 오전/오후를 그 시각 기준으로 맞춘다.
        global RUN_STAMP
        RUN_STAMP = f"{end:%Y-%m-%d %H:%M}"

    log("=" * 68)
    기간 = f"{hours}시간" if hours < 48 else f"{hours / 24:.0f}일"
    log(f" 나라장터 공고 수집   {begin:%Y-%m-%d} ~ {end:%Y-%m-%d}  ({기간})")
    if 기준_raw:
        log(f" [재현] {end:%Y-%m-%d %H:%M} 에 돌렸다고 치고 그 시각까지만 봅니다."
            f" ({begin:%m/%d %H:%M} ~ {end:%m/%d %H:%M})")
    낮은컷 = (f"(연기금·공제회 {sc.cut_review_low}점)"
             if sc.institutions_low_cut and sc.cut_review_low < sc.cut_review else "")
    log(f" 업종코드 {', '.join(cfg['industry_codes'])}"
        f" · 수집 {sc.cut_collect}점 · 검토 {sc.cut_review}점{낮은컷}"
        f" · PDF검사 {'켬' if use_pdf else '끔'}"
        f" · 메일 {'켬' if cfg['mail']['enabled'] else '끔'}"
        f"{' · 목록만' if list_only else ''}")
    log(f" 저장 위치 : {root}")
    log("=" * 68)

    with RunLock(root):
        return run(cfg, sc, root, begin, end, use_pdf, list_only)


def run(cfg: dict, sc: Scorer, root: Path, begin: datetime, end: datetime,
        use_pdf: bool, list_only: bool) -> int:
    seen_path = root / SEEN_NAME
    seen = load_seen(seen_path)

    log("\n[1차] 면허제한 조회 중... (양이 많아 몇 분 걸립니다)")
    lic_index, lic_빠짐 = build_license_index(cfg["service_key"], begin, end, root)
    n_hit = sum(1 for c, _ in lic_index.values() if c & set(cfg["industry_codes"]))
    log(f"  면허제한 있는 공고 {len(lic_index):,}건 · 그중 업종코드 일치 {n_hit:,}건")
    # 면허제한은 A 판정의 근거 하나일 뿐이라 이것만 끊겨도 수집은 돌아간다.
    # 다만 A 를 놓칠 수 있으므로 회차는 '일부실패' 로 본다 (필수=False 는
    # '이것만 실패해도 완전실패는 아니다' 라는 뜻이다).
    조회기록 = [check("면허제한", len(lic_index), lic_빠짐, 필수=False)]
    if lic_빠짐:
        log("  [경고] 조회가 중간에 끊겨 일부만 받았습니다. A등급을 놓칠 수 있습니다.")
        log("         받아둔 구간은 캐시에 남았습니다. 같은 명령으로 다시 실행하면")
        log("         못 받은 구간부터 이어받습니다.")

    # 수집분과 검토 필요분을 한 목록에 담는다 (요구사항_v3 회신 3번).
    결과: list[list] = []
    처리됨: set = set()
    차수기록: dict[str, int] = {}   # 공고번호 → 이번에 본 차수 (이력에 남긴다)
    전체 = 0
    pdf_검사수 = 0
    저장된_pdf: list[Path] = []

    for target in cfg["targets"]:
        log(f"\n[2차] {target} 조회 중...")
        raws, 빠짐 = call_api(ENDPOINTS[target], cfg["service_key"], begin, end,
                             label=target)
        조회기록.append(check(target, len(raws), 빠짐))
        if 빠짐:
            log(f"  [경고] {target} 조회가 완전하지 않습니다: {', '.join(빠짐)}")
        전체 += len(raws)

        live, 취소목록, n_옛차수 = pick_live(raws)

        # 이미 보낸 공고는 건너뛴다. 다만 '그때와 값이 같을 때만' 이다.
        # 마감이 미뤄졌거나 예산이 바뀌었으면 다시 알려야 한다.
        신규, 변경 = [], []
        for r in live:
            uid = no_of(r)
            if uid in 처리됨:
                continue
            처리됨.add(uid)
            차수기록[uid] = notice_ord(r)
            이전 = seen.get(이력키(SRC_G2B, uid))
            if 이전 is None:
                신규.append(r)
                continue
            내역 = diff_notice(이전, r, notice_ord(r))
            if 내역:
                변경.append((r, 내역))
            elif not 이전:
                # 옛 이력이라 견줄 값이 없다. 지금 값을 채워 다음 회차부터
                # 비교되게 한다. 이번에는 알리지 않는다.
                seen[이력키(SRC_G2B, uid)] = first_snapshot(r)

        # 우리가 보낸 공고가 취소됐으면 그것도 알린다.
        취소행 = []
        for uid, r in 취소목록.items():
            이전 = seen.get(이력키(SRC_G2B, uid))
            if 이전:                      # 빈 dict(옛 이력)면 등급을 모른다
                취소행.append(cancel_row(이전, r))
        결과 += 취소행

        log(f"  {len(raws):,}건 조회 → 취소 {len(취소목록):,}건"
            f"·옛차수 {n_옛차수:,}건 제외 → 신규 {len(신규):,}건")
        if 변경 or 취소행:
            log(f"  이미 보낸 공고 중 변경 {len(변경)}건 · 취소 {len(취소행)}건")
        for row in 취소행:
            log(f"    [취소] {row[COL_DM] or row[COL_NT]} — {row[COL_TITLE][:44]}")

        # PDF 를 열기 전에 점수로 후보를 먼저 좁힌다. 이게 없으면 3만 건을 다 연다.
        후보 = []
        for r in 신규:
            uid = (r.get("bidNtceNo") or "").strip()
            codes, _ = lic_index.get(uid, (set(), ""))
            if codes & set(cfg["industry_codes"]):
                후보.append(r)
                continue
            orgs = f"{r.get('dminsttNm') or ''} {r.get('ntceInsttNm') or ''}"
            pts, _ = sc.score((r.get("bidNtceNm") or ""), orgs)
            if pts >= sc.review_cut(orgs):
                후보.append(r)
        # 변경된 건은 점수와 무관하게 다시 본다. 이미 한 번 보낸 공고라
        # 지금 점수가 문턱 아래로 내려갔더라도 '그 공고가 바뀌었다' 는
        # 알려야 한다. 판정이 안 나오면 지난번 등급을 그대로 쓴다.
        후보 += [r for r, _ in 변경]
        변경내역 = {no_of(r): 내역 for r, 내역 in 변경}

        if use_pdf and 후보:
            log(f"  [3차] 후보 {len(후보)}건의 첨부 PDF 를 열어봅니다...")
            pdf_검사수 += len(후보)

        for r in 후보:
            내역 = 변경내역.get(no_of(r), "")
            이전 = seen.get(이력키(SRC_G2B, no_of(r))) or {}
            verdict = judge(r, cfg, sc, lic_index, use_pdf)
            if verdict is None:
                if not 내역:
                    continue
                # 바뀐 건인데 지금 기준으로는 안 걸린다. 지난번 판정을 쓴다.
                grade = 이전.get("정확도") or "C"
                pts = 이전.get("점수") or 0
                kw, why, pdf_body = "-", "지난 회차 판정을 그대로 씁니다", None
            else:
                grade, pts, kw, why, pdf_body = verdict

            title = (r.get("bidNtceNm") or "").strip()
            dm = (r.get("dminsttNm") or "").strip()
            nt = (r.get("ntceInsttNm") or "").strip()
            no = (r.get("bidNtceNo") or "").strip()
            n_files = sum(1 for i in range(1, 11)
                          if (r.get(f"ntceSpecDocUrl{i}") or "").strip())

            # 변경으로 다시 올라온 건은 API 가 뭐라 하든 '변경' 으로 적는다.
            종류 = "변경" if 내역 else notice_kind(r)
            row = [grade, 종류, 내역, (r.get("bidNtceDt") or "")[:10],
                   dm, nt, title, deadline(r),
                   money(r.get("asignBdgtAmt")), money(r.get("presmptPrce")),
                   target, pts, kw, why,
                   n_files, SRC_G2B, no, (r.get("bidNtceDtlUrl") or "").strip(),
                   RUN_STAMP]

            결과.append(row)

            # 엑셀에 오른 건이면 등급과 무관하게 공고문을 남긴다.
            # 특히 C등급(검토 필요)은 사람이 열어봐야 하는 것이라 더 필요하다.
            org = dm or nt or "기관미상"
            log(f"    [{grade}]{' [변경]' if 내역 else ''} {org} — {title[:44]}")
            if 내역:
                log(f"           {내역}")
            if not list_only and pdf_body:
                saved = save_pdf(pdf_body, org, title, no, root)
                if saved:
                    저장된_pdf.append(saved)

    # 처리한 것만 이력에 남긴다. 조회했지만 점수 미달로 버린 건은
    # 나중에 점수표를 고쳤을 때 다시 판정할 수 있어야 하므로 남기지 않는다.
    #
    # 번호만이 아니라 지금 값을 통째로 남긴다. 다음 회차에 이 값과 견줘
    # 무엇이 바뀌었는지 말하기 위해서다.
    for row in 결과:
        키 = 이력키(str(row[COL_SRC] or SRC_G2B), str(row[COL_NO]))
        if row[COL_KIND] == "취소":
            # 취소된 공고는 다시 살아나지 않는다. 이력에서 뺀다. 그대로 두면
            # 같은 번호가 조회될 때마다 취소 알림이 되풀이된다.
            seen.pop(키, None)
            continue
        seen[키] = snapshot(row, 차수기록.get(str(row[COL_NO]), 0))

    # 등급순(S→C), 같은 등급이면 점수 높은 것부터. 엑셀도 메일도 이 순서다.
    결과.sort(key=lambda r: (GRADE_ORDER.get(r[COL_GRADE], 9), -int(r[COL_SCORE])))

    saved_ok = append_rows(root / COLLECT_XLSX, 결과)
    if saved_ok:
        save_seen(seen_path, seen)

    n_검토 = sum(1 for r in 결과 if r[COL_GRADE] == "C")
    log("\n" + "-" * 68)
    log(f" 전체 {전체:,}건 조회 → 수집 {len(결과) - n_검토}건 · 검토 필요 {n_검토}건")
    if use_pdf:
        log(f" PDF 검사 {pdf_검사수}건 · 공고문 저장 {len(저장된_pdf)}건")
    for g in ("A", "B", "C"):
        n = sum(1 for r in 결과 if r[COL_GRADE] == g)
        if n:
            log(f"   {g} {n:3d}건  ({GRADE_DESC[g]})")
    if saved_ok:
        log(f" 공고목록 : {root / COLLECT_XLSX}")
    if 저장된_pdf:
        log(f" 공고문   : {root / PDF_DIR}")
    log("-" * 68)

    old_review = root / OLD_REVIEW_XLSX
    if old_review.exists():
        log(f"\n[안내] 이제 검토 필요분(C)도 {COLLECT_XLSX} 에 함께 들어갑니다.")
        log(f"       예전 {OLD_REVIEW_XLSX} 은 더 쓰지 않으니 옮겨두거나 지우세요.")

    상태 = run_status(조회기록)
    빠진구간 = missing_phrase(조회기록)
    save_check_log(root, 조회기록, 상태, len(결과), begin, end)

    if 상태 != 정상:
        log(f"\n[조회 상태] {상태} — 못 받은 구간: {빠진구간}")
        log(f"  다시 돌리려면:  {rerun_command(begin, end)}")

    if 상태 == 완전실패:
        # 결과 메일을 보내면 '오늘은 공고가 없었다' 로 읽힌다. 그것만은 막는다.
        log("  [경고] 목록 조회가 모두 실패해 결과 메일을 보내지 않습니다.")
        notify_error(cfg, 조회기록, root, begin, end)
    else:
        notify(cfg, 결과, root, end, period_phrase(begin, end),
               빠진구간, 상태)

    if not 결과:
        log("\n조건에 맞는 공고가 없습니다.")
        log("점수표를 확인하려면:  " + ("편집기의 '점수표' 탭에서 검증 버튼" if FROZEN
                                  else "python 점수표_검증.py"))
    return 0


if __name__ == "__main__":
    code = 1
    try:
        code = main()
    except KeyboardInterrupt:
        log("\n중단했습니다.")
    except SystemExit as exc:
        if isinstance(exc.code, str):
            log(f"\n{exc.code}")
        code = 0 if exc.code in (0, None) else 1
    except Exception as exc:  # noqa: BLE001
        log(f"\n[예기치 못한 오류] {type(exc).__name__}: {exc}")
    raise SystemExit(code)
